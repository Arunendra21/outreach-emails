from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

companies = [
    (1,"Blink22","blink22.com","Web Dev","hello@blink22.com","contact","","","","","Medium","Send cold email; ask to forward to hiring manager"),
    (2,"Shakuro","shakuro.com","Web/UX","hi@shakuro.com","contact","","","","","Medium","Send cold email"),
    (3,"ExpandX","expandx.ae","Web Dev","info@expandx.ae","contact","","","","","High","Dubai-based"),
    (4,"Digital Gravity","digitalgravity.ae","Web Dev","discover@digitalgravity.ae","contact","","","","","High","Dubai-based; email scraped from contact page ✅"),
    (5,"Branex","branex.ae","Web Design","info@branex.ae","contact","","","","","High","Dubai-based"),
    (6,"BI Communications","bicommunications.ae","Web Dev","info@bicommunications.ae","contact","","","","","High","Dubai-based"),
    (7,"Synergy Labs","synergylabs.co","App Dev","hi@synergylabs.co","contact","","","","","Medium","Send cold email"),
    (8,"7EDGE","7edge.com","Software Dev","hello@7edge.com","contact","","","","","Medium","Send cold email"),
    (9,"GoodCore Software","goodcoresoft.com","Software Dev","hello@goodcoresoft.com","contact","","","","","Medium","Send cold email"),
    (10,"Net Solutions","netsolutions.com","Web Dev","hello@netsolutions.com","contact","","","","","Medium","Send cold email"),
    (11,"Webdew","webdew.com","Web Dev","info@webdew.com","contact","","","","","Medium","Send cold email"),
    (12,"Intelivita","intelivita.com","App Dev","hello@intelivita.com","contact","","","","","Medium","Send cold email"),
    (13,"Finoit Technologies","finoit.com","SaaS Dev","hello@finoit.com","contact","","","","","Medium","Send cold email"),
    (14,"Simform","simform.com","Engineering","business@simform.com","contact","","","","","Medium","Send cold email"),
    (15,"TechAhead","techaheadcorp.com","App Dev","info@techaheadcorp.com","contact","","","","","Medium","Send cold email"),
    (16,"Appinventiv","appinventiv.com","App Dev","career@appinventiv.com","careers","","","","","Medium","DIRECT CAREER EMAIL scraped ✅"),
    (17,"Mindinventory","mindinventory.com","App Dev","info@mindinventory.com","contact","","","","","Medium","Send cold email"),
    (18,"Bacancy Technology","bacancytechnology.com","Web Dev","info@bacancytechnology.com","contact","","","","","Medium","Send cold email"),
    (19,"Cubix","cubix.co","App/Game Dev","hello@cubix.co","contact","","","","","Medium","Send cold email"),
    (20,"SoluLab","solulab.com","Blockchain/Web","hello@solulab.com","contact","","","","","Medium","Send cold email"),
    (21,"Apptunix","apptunix.com","App Dev","info@apptunix.com","contact","","","","","Medium","Send cold email"),
    (22,"Matellio","matellio.com","Software Dev","info@matellio.com","contact","","","","","Medium","Send cold email"),
    (23,"IndiaNIC","indianic.com","App Dev","hello@indianic.com","contact","","","","","Medium","Send cold email"),
    (24,"OpenXcell","openxcell.com","App Dev","contact@openxcell.com","contact","","","","","Medium","Send cold email"),
    (25,"Brainvire","brainvire.com","Web/Cloud","info@brainvire.com","contact","","","","","Medium","Send cold email"),
    (26,"Zealous System","zealousys.com","App Dev","bd@zealousys.com","contact","","","","","Medium","BD team"),
    (27,"Hidden Brains","hiddenbrains.com","App Dev","hr@hiddenbrains.com","careers","","","","","Medium","DIRECT HR EMAIL scraped ✅"),
    (28,"ValueCoders","valuecoders.com","Outsource Dev","sales@valuecoders.com","contact","","","","","Medium","Sales team"),
    (29,"ToXSL Technologies","toxsl.com","Web Dev","sales@toxsl.com","contact","","","","","Medium","Send cold email"),
    (30,"Successive Technologies","successive.tech","Cloud Dev","hello@successive.tech","contact","","","","","Medium","Send cold email"),
    (31,"Rishabh Software","rishabhsoft.com","Enterprise Dev","info@rishabhsoft.com","contact","","","","","Medium","Send cold email"),
    (32,"Narola Infotech","narolainfotech.com","Web Dev","inquiry@narolainfotech.com","contact","Raj","—","raj@narolainfotech.com","","Medium","Named person scraped from contact page ✅"),
    (33,"Techuz","techuz.com","App Dev","hello@techuz.com","contact","","","","","Medium","Send cold email"),
    (34,"Clavax Technologies","clavax.com","App Dev","hr@clavax.com","careers","","","","","Medium","DIRECT HR EMAIL scraped ✅ — also sales@clavax.com"),
    (35,"Vinfotech","vinfotech.com","Sports Dev","info@vinfotech.com","contact","","","","","Medium","Send cold email"),
    (36,"Algoworks","algoworks.com","Salesforce/Web","info@algoworks.com","contact","","","","","Medium","Send cold email"),
    (37,"Quytech","quytech.com","Blockchain/AI","hr@quytech.com","careers","","","","","Medium","DIRECT HR EMAIL scraped ✅ — also sales@quytech.com"),
    (38,"Astra Tech","astratech.ae","AI/Fintech","info@astratech.ae","contact","","","","","High","Dubai AI company"),
    (39,"ai71","ai71.ai","AI Products","contact@ai71.ai","contact","","","","","High","UAE AI company; contact@ scraped ✅"),
    (40,"Plavno","plavno.io","AI/Web Dev","hello@plavno.io","contact","","","","","Medium","Send cold email"),
    (41,"Sarwa","sarwa.co","Fintech","hello@sarwa.co","contact","","","","careers.sarwa.co","High","Dubai fintech"),
    (42,"Beehive","beehive.ae","Fintech","hello@beehive.ae","contact","","","","","High","Dubai fintech"),
    (43,"NOW Money","nowmoney.me","Fintech","hello@nowmoney.me","contact","","","","","High","Dubai fintech"),
    (44,"Tabby","tabby.ai","BNPL","hello@tabby.ai","contact","","","","tabby.ai/en-AE/careers","High","Apply via portal + cold email"),
    (45,"Ziina","ziina.com","Payments","partnerships@ziina.com","contact","","","","","High","partnerships@ scraped from about page ✅"),
    (46,"Stake","getstake.com","PropTech","hello@getstake.com","contact","","","","","High","Dubai PropTech"),
    (47,"PayTabs","paytabs.com","Payments","uaesales@paytabs.com","contact","","","","","High","UAE payments; uaesales@ + customercare@ scraped ✅"),
    (48,"Telr","telr.com","Payments","sales@telr.com","contact","","","","","High","Multiple emails scraped ✅: sales@, support@, partner@, marketing@telr.com"),
    (49,"Tarabut Gateway","tarabut.com","Open Banking","hello@tarabut.com","contact","","","","","High","UAE Open Banking"),
    (50,"Lean Technologies","leantech.me","Open Banking","hello@leantech.me","contact","","","","","High","UAE Open Banking"),
    (51,"Property Finder","propertyfinder.ae","PropTech","careers@propertyfinder.ae","careers","","","","","High","Direct HR inbox"),
    (52,"Dubizzle","dubizzle.com","Classifieds","careers@dubizzle.com","careers","","","","","High","Direct HR inbox"),
    (53,"Bayut","bayut.com","PropTech","careers@bayut.com","careers","Mirna Al Sayegh","Sr. TA Specialist","m***@bayut.com (use Apollo)","https://ae.linkedin.com/in/mirna-al-sayegh-assoc-cipd-72baaa42","High","Email careers@ AND DM Mirna on LinkedIn"),
    (54,"Huspy","huspy.com","PropTech","hello@huspy.com","contact","John Michael Razon","People Team","john.razon@huspy.io","https://ae.linkedin.com/in/john-michael-razon-ba249066","High","VERIFIED - email directly"),
    ("54b","Huspy (2nd contact)","huspy.com","PropTech","hello@huspy.com","contact","Grasielly D.","People Lead","grasielly@huspy.io (est)","https://ae.linkedin.com/in/grasielly-d-06136292","High","DM LinkedIn to confirm"),
    (55,"Ajar Online","ajaronline.com","PropTech","hello@ajaronline.com","contact","","","","","High","Dubai PropTech"),
    (56,"Vezeeta","vezeeta.com","HealthTech","hello@vezeeta.com","contact","","","","","High","UAE HealthTech"),
    (57,"Okadoc","okadoc.com","HealthTech","hello@okadoc.com","contact","","","","","High","Dubai HealthTech"),
    (58,"Altibbi","altibbi.com","HealthTech","info@altibbi.com","contact","","","","","High","UAE HealthTech"),
    (59,"Bayzat","bayzat.com","HR Tech","hello@bayzat.com","contact","Sanjo Joshi","Sr. Talent Resourcer","sanjo.joshi@bayzat.com (est)","","High","DM LinkedIn to confirm"),
    (60,"Rise","rise.ae","HR Tech","hello@rise.ae","contact","","","","","High","Dubai HR Tech"),
    (61,"Fetchr","fetchr.us","Logistics","hello@fetchr.us","contact","","","","","Medium","Verify active - domain may redirect"),
    (62,"Cafu","cafu.com","Logistics","hello@cafu.com","contact","Pranita Talukdar","HR Manager","p***@cafu.com (use Apollo)","","High","DM LinkedIn to get full email"),
    (63,"Trukkin","trukkin.com","Logistics","info@trukkin.com","contact","","","","","High","Dubai logistics"),
    (64,"Swvl","swvl.com","Transit Tech","hello@swvl.com","contact","","","","","Medium","Verify if active"),
    (65,"Anghami","anghami.com","Music Tech","hello@anghami.com","contact","","","","","High","UAE music tech"),
    (66,"Yalla","yalla.live","Social/Gaming","bd@yalla.live","contact","","","","","High","Dubai gaming"),
    (67,"Kitopi","kitopi.com","FoodTech","hello@kitopi.com","contact","","","","","High","Dubai cloud kitchens"),
    (68,"Entertainer","theentertainerme.com","Loyalty Tech","hello@theentertainerme.com","contact","","","","","High","Dubai loyalty tech"),
    (69,"Cofe App","cofeapp.com","Retail Tech","hello@cofeapp.com","contact","","","","","High","Dubai retail tech"),
    (70,"Bayt.com","bayt.com","Job Platform","support@bayt.com","contact","","","","","High","UAE job platform"),
    (71,"Careem","careem.com","SuperApp","careers@careem.com","careers","","","","jobs.careem.com","High","Direct HR careers email"),
    (72,"Noon","noon.com","eCommerce","careers@noon.com","careers","","","","","High","Confirmed direct careers email"),
    (73,"Talabat","talabat.com","Food Delivery","inquiries@talabat.com","careers","","","","careers.deliveryhero.com/talabat","High","Use inquiries@ + apply via portal"),
    (74,"Saffron Tech","saffrontech.net","Web Dev","info@saffrontech.net","contact","","","","","Medium","Send cold email"),
    (75,"Itransition","itransition.com","Enterprise Dev","info@itransition.com","contact","","","","","Medium","Send cold email"),
    (76,"instinctools","instinctools.com","SaaS Dev","hello@instinctools.com","contact","","","","","Medium","Send cold email"),
    (77,"SDLC Corp","sdlccorp.com","Software Dev","hello@sdlccorp.com","contact","","","","","Medium","Send cold email"),
    (78,"G42","g42.ai","AI/Cloud","contact@g42.ai","contact","","","","careers.g42.ai","High","UAE AI giant - apply via portal too"),
    (79,"Presight","presight.ai","AI/Analytics","info@presight.ai","contact","","","","","High","Abu Dhabi AI analytics"),
    (80,"Pure Harvest","pure-harvest.com","AgriTech","hello@pure-harvest.com","contact","","","","","High","UAE AgriTech"),
    (81,"Coda Payments","codapayments.com","Gaming/Payments","hello@codapayments.com","contact","","","","","Medium","Send cold email"),
    (82,"Workruit","workruit.com","HR Tech","hello@workruit.com","contact","","","","","Medium","Send cold email"),
    (83,"d1g1t","d1g1t.com","Fintech","info@d1g1t.com","contact","","","","","Medium","Send cold email"),
    (84,"Network International","networkinternational.ae","Payments","info@networkinternational.ae","contact","","","","","High","UAE payments giant"),
    (85,"Magnati","magnati.com","Payments","info@magnati.com","contact","","","","","High","UAE payments"),
    (86,"Elmenus","elmenus.com","FoodTech","hello@elmenus.com","contact","","","","","Medium","Send cold email"),
    (87,"GoBOLT","gobolt.com","Logistics","hello@gobolt.com","contact","","","","","Medium","Send cold email"),
    (88,"Etisalat Digital (e&)","eand.com","Telecom/Tech","info@eand.com","contact","","","","","High","Major UAE telecom tech arm"),
    # ── NEW COMPANIES (scraped round 2) ──────────────────────────────────────
    (89,"TekRevol","tekrevol.com","App/Web Dev","info@tekrevol.com","contact","","","","","High","Dubai-based; email confirmed"),
    (90,"PLAN A Technology","plana.tech","Software Dev","contact@plana.tech","contact","","","","","High","Dubai Silicon Oasis; email confirmed"),
    (91,"Team Rhino","teamrhino.ae","Web Dev","hello@teamrhinoltd.com","contact","","","","","High","Sheikh Zayed Rd Dubai; email confirmed"),
    (92,"Adapts Media","adaptsmedia.com","Digital Marketing","info@adaptsmedia.com","contact","Ankita","Manager","ankita@adaptsmedia.com","","High","Named person scraped from contact page ✅"),
    (93,"Alaan","alaan.com","Fintech/Expense Mgmt","careers@alaanpay.com","careers","","","","","High","VERIFIED careers email - send directly"),
    (94,"Qashio","qashio.com","Fintech/Spend Mgmt","hello@qashio.com","contact","","","","careers.qashio.com","High","Dubai fintech; portal + cold email"),
    (95,"NymCard","nymcard.com","Embedded Finance","careers@nymcard.com","careers","","","","","High","DIRECT CAREER EMAIL scraped ✅ — also contact@, press@nymcard.com"),
    (96,"Baraka","baraka.io","Investment Tech","hello@baraka.io","contact","","","","","High","Dubai investment app"),
    (97,"BitOasis","bitoasis.net","Crypto/Blockchain","info@bitoasis.net","contact","","","","","High","Dubai crypto exchange"),
    (98,"Alef Education","alef.ae","EdTech/AI","info@alef.ae","contact","","","","","High","Dubai AI-powered EdTech"),
    (99,"Help AG","helpag.com","Cybersecurity","info@helpag.com","contact","","","","","High","Dubai cyber defense; part of e&"),
    (100,"Digital14","digital14.com","Digital/Cyber","info@digital14.com","contact","","","","","High","Abu Dhabi digital trust"),
    (101,"BRAVA 360 Digital","brava360.digital","Digital Agency","info@brava360.digital","contact","","","","","High","Dubai digital agency"),
    (102,"Digital Hub Sol","digitalhubsol.ae","Digital Marketing","info@digitalhubsol.ae","contact","","","","","High","Dubai digital agency"),
    (103,"Magneto IT Solutions","magnetoitsolutions.com","eCommerce Dev","info@magnetoitsolutions.com","contact","","","","","Medium","Send cold email"),
    (104,"BitsWits","bitswits.co","App Dev","info@bitswits.co","contact","","","","","Medium","Send cold email"),
    (105,"Soharon Infotech","soharon.com","App Dev","info@soharon.com","contact","","","","","Medium","Send cold email"),
    (106,"Hexagon IT Solutions","hexagonitsolutions.com","Software Dev","info@hexagonitsolutions.com","contact","","","","","Medium","Send cold email"),
    (107,"Computools","computools.com","Software Dev","info@computools.com","contact","","","","","Medium","Send cold email"),
    (108,"CraftedQ","craftedq.com","Design/Dev","info@craftedq.com","contact","","","","","Medium","Send cold email"),
    (109,"Dot IT","dotit.org","Web Dev","info@dotit.org","contact","","","","","High","Dubai web dev agency"),
    (110,"SEO Tech Experts","seotechexperts.ae","Digital Marketing","info@seotechexperts.ae","contact","","","","","High","Dubai-based .ae domain"),
    (111,"KONSTANT INFOSOLUTIONS","konstantinfo.com","App Dev","info@konstantinfo.com","contact","","","","","Medium","Send cold email"),
    (112,"Kuchoriya TechSoft","kuchoriyatechsoft.com","App Dev","info@kuchoriyatechsoft.com","contact","","","","","Medium","Send cold email"),
    (113,"Quixta","quixta.com","Web Dev","info@quixta.com","contact","","","","","Medium","Send cold email"),
    (114,"GCC Marketing","gcc-marketing.com","Digital Marketing","info@gcc-marketing.com","contact","","","","","High","Dubai digital marketing"),
    (115,"Incrementors","incrementors.com","Digital Marketing","info@incrementors.com","contact","","","","","Medium","Send cold email"),
    (116,"Classera","classera.com","EdTech","info@classera.com","contact","","","","","High","Saudi/UAE EdTech platform"),
    (117,"Glimpse","theglimpseproject.com","Creative Tech","info@theglimpseproject.com","contact","","","","","High","Dubai creative agency"),
    (118,"Suffescom Solutions","suffescom.com","Blockchain/Web","info@suffescom.com","contact","","","","","Medium","Send cold email"),
    (119,"Probey Services","probeyservices.com","Web Dev","info@probeyservices.com","contact","","","","","Medium","Send cold email"),
    (120,"EffectiveSoft","effectivesoft.com","Enterprise Dev","info@effectivesoft.com","contact","","","","","Medium","Send cold email"),
    (121,"Enova by Veolia","enova-me.com","Facility/Energy Tech","recruitment@enova-me.com","careers","","","","","High","VERIFIED from LinkedIn post — direct recruitment inbox"),
    (122,"Melodica UAE","melodica.ae","EdTech/Music","sasha.w@melodica.ae","careers","Sasha W.","Hiring Manager","sasha.w@melodica.ae","","High","VERIFIED from LinkedIn post — named contact"),
    (123,"Deriv","deriv.com","Fintech/Trading","recruitment@deriv.com","careers","Hesilin Pouzi","TA Specialist","hesilin.pouzi@deriv.com (est)","https://www.linkedin.com/in/hesilinpouzi","High","Dubai office; TA Specialist found on LinkedIn"),
    (124,"Emirates Group","careers.emirates.com","Aviation/Tech","careers@emirates.com","careers","","","","careers.emirates.com","High","Paid internship programme — 4000 AED/month"),
    (125,"Mubadala Investment","mubadala.com","Investment/Tech","careers@mubadala.com","careers","","","","mubadala.com/en/careers","High","Abu Dhabi sovereign fund — internship programme confirmed"),
    (126,"BCG X Dubai","bcg.com","Consulting/AI","careers-uae@bcg.com","careers","","","","careers.bcg.com","High","CONFIRMED AI Engineer + Data Scientist internships in Dubai 2026"),
    (127,"Thales UAE","thalesgroup.com","Defense/Tech","careers-uae@thalesgroup.com","careers","","","","careers.thalesgroup.com","High","CONFIRMED AI/ML intern roles in Dubai/Abu Dhabi 2026"),
    (128,"Carmatec","carmatec.com","Software Dev","careers@carmatec.com","careers","","","teamhr@carmatec.com","","Medium","careers@ + teamhr@ both confirmed from search"),
    (129,"SISGAIN","sisgain.ae","Software Dev","hr@sisgain.ae","careers","","","","","High","DIRECT HR EMAIL scraped ✅ — also hello@sisgain.ae"),
    (130,"AIQU Solutions","aiqusearch.com","AI/Recruitment Tech","hello@aiqusearch.com","contact","","","","","High","Abu Dhabi/Dubai/Riyadh — email confirmed"),
    (131,"Quickwork","quickwork.co","Integration/SaaS","info@quickwork.co","contact","","","","","Medium","Send cold email"),
    (132,"Hyperlink InfoSystem","hyperlinkinfosystem.com","App Dev","info@hyperlinkinfosystem.com","contact","","","","","Medium","Send cold email"),
    (133,"Xicom Technologies","xicom.biz","Software Dev","info@xicom.biz","contact","","","","","Medium","Dubai Business Bay office"),
    (134,"CodeNinja","codeninja.ae","App Dev","info@codeninja.ae","contact","","","","","High","Dubai-based .ae domain"),
    (135,"iCreativez Technology","icreativez.com","Web Dev","info@icreativez.com","contact","","","","","Medium","Dubai + Pakistan + Qatar"),
    (136,"Careem (BCG spinout-style)","careemforwork.com","CorpTech","careers@careemforwork.com","careers","","","","","High","Careem B2B arm — separate internship pipeline"),
    (137,"Talabat Tech","tech.talabat.com","FoodTech/Engineering","tech@talabat.com","contact","","","","careers.deliveryhero.com/talabat","High","Talabat confirmed AI Engineering Intern 2026"),
    (138,"Phaedra Solutions","phaedrasolutions.com","Software Dev","hello@phaedrasolutions.com","contact","","","","","High","Dubai + UK + USA"),
    (139,"Vision Technologies","visiontechsys.com","IT Services","info@visiontechsys.com","contact","","","","","High","Sharjah UAE — 300+ employees"),
    (140,"Exceed IT Services","exceedgulf.com","Enterprise IT","info@exceedgulf.com","contact","","","","","High","Abu Dhabi UAE"),
    # ── ROUND 3: LinkedIn browser scrape ─────────────────────────────────────
    (141,"SQUATWOLF","squatwolf.com","FitTech/eComm","careers@squatwolf.com","careers","","","","https://www.linkedin.com/company/squat-wolf/","High","Dubai fitness tech; support@ confirmed, careers@ standard pattern"),
    (142,"Deriv","deriv.com","Fintech/Trading","careers@deriv.com","careers","","","","careers.deriv.com","High","ACTIVE AI hiring confirmed — LLM, agent systems, fintech"),
    (143,"Huspy (3rd contact — CEO)","huspy.com","PropTech","CEO@huspy.com","contact","Tariq Tber","CEO / Co-Founder","tariq@huspy.io (pattern)","https://www.linkedin.com/in/tariqtber/","High","Founder-led — cold email to CEO gets replied to at startups"),
    # ── ROUND 4: Gandhinagar / GIFT City companies ───────────────────────────
    (144,"DRC Systems","drcsystems.com","Software Dev / IT Consulting","careers@drcsystems.com","careers","","","","https://careers.drcsystems.com","High","VERIFIED ✅ careers@ + info@ + sales@; GIFT City 24th Floor; Dubai office too"),
    (145,"Argusoft India","argusoft.com","Healthcare IT / Software","jobs@argusoft.com","careers","","","","https://careers.argusoft.com","High","DIRECT JOBS EMAIL ✅ — A66 GIDC Sector-25 Gandhinagar"),
    (146,"Odoo India","odoo.com","ERP / SaaS","—","careers","","","","https://www.odoo.com/jobs","High","IT Tower 3 Infocity Gandhinagar; ₹20K/month intern stipend; apply via portal"),
    (147,"Bosc Tech Labs","bosctechlabs.com","App Dev / AI","hr@bosctechlabs.com","careers","","","","","High","DIRECT HR EMAIL ✅ — Sargasan Gandhinagar; also contact@bosctechlabs.com"),
    (148,"Cybage Software","cybage.com","Enterprise Dev","careers@cybage.com","careers","","","","https://www.cybage.com/careers","High","DIRECT CAREERS EMAIL ✅ — Tower II Infocity Gandhinagar"),
    (149,"Prismetric Technologies","prismetric.com","App Dev / Digital","biz@prismetric.com","contact","","","","https://www.prismetric.com/job-opportunities/","High","604 IT Tower 1, Infocity Gandhinagar 382007"),
    (150,"Decimal Point Analytics","decimalpointanalytics.com","Data Analytics / AI","info@decimalpointanalytics.com","contact","","","","","High","D-601 WTC GIFT City Gandhinagar; financial research & ML/AI"),
    (151,"Infibeam Avenues (AvenuesAI)","avenuesai.com","Digital Payments / Fintech","contactus@avenuesai.com","contact","","","","https://www.avenuesai.com/career-opportunities","High","VERIFIED ✅ contactus@ + ir@ + corpcom@; 28th Floor GIFT Two, GIFT City"),
    (152,"Prometheanz","prometheanz.com","Enterprise Software","contact@prometheanz.com","contact","","","","","High","Office 120, Infocity Tower-1 Gandhinagar 382007"),
    (153,"Oddeven Infotech","oddeveninfotech.com","Salesforce / Web Dev","hello@oddeveninfotech.com","contact","","","","","High","Infocity Gandhinagar; Salesforce, AI, digital transformation"),
    (154,"Warlock Technologies","warlocktechnologies.com","Odoo / ERP","info@warlocktechnologies.com","contact","","","","","High","Kudasan Gandhinagar; Official Odoo Partner"),
    (155,"Zenkins Technologies","zenkins.com","Web/Mobile Dev","careers@zenkins.com","careers","","","","","Medium","DIRECT CAREERS EMAIL ✅ — also contact@zenkins.com; near GIFT City"),
    (156,"ManekTech","manektech.com","App Dev / Web Dev","jobs@manektech.com","careers","","","","","Medium","DIRECT JOBS EMAIL ✅ — also info@manektech.com; 450+ employees"),
    (157,"Anblicks","anblicks.com","Data / AI / Analytics","careers@anblicks.com","careers","","","","https://anblicks.keka.com/careers/","High","DIRECT CAREERS EMAIL ✅ — Enterprise AI; Gandhinagar office"),
    (158,"AxisTechnoLabs","axistechnolabs.com","Odoo / Python Dev","career@axistechnolabs.com","careers","","","","","Medium","DIRECT CAREER EMAIL ✅ — Python internships; near Gandhinagar"),
    (159,"Decentro","decentro.tech","Fintech Infrastructure","hello@decentro.tech","contact","","","","https://decentro.tech/careers","High","GIFT City subsidiary (DECFIN); payment APIs; 1600+ customers"),
    (160,"Gujarat Informatics Limited","gil.gujarat.gov.in","e-Governance / IT","infogil@gujarat.gov.in","contact","","","","https://gil.gujarat.gov.in/Careers","High","Govt IT arm; Karmayogi Bhavan Sector-10A Gandhinagar"),
    (161,"AgroStar","corporate.agrostar.in","AgriTech / AI","—","careers","","","","https://corporate.agrostar.in/contact","High","5th Floor Infocity Tower-1 Gandhinagar; AI-powered agri platform"),
    (162,"Accenture GIFT City","accenture.com","IT Consulting / AI","—","careers","","","","https://www.accenture.com/in-en/careers","High","GIFT City office; cybersecurity & IT consulting"),
    (163,"Zoho GIFT City","zoho.com","SaaS / ERP / CRM","—","careers","","","","https://www.zoho.com/careers.html","High","GIFT City presence; cloud-based ERP/CRM; apply via portal"),
    (164,"Razorpay GIFT City","razorpay.com","Fintech / Payments","—","careers","","","","https://razorpay.com/jobs/","High","GIFT City cross-border payments center; apply via portal"),
    (165,"LTIMindtree GIFT City","ltimindtree.com","IT Services / AI","—","careers","","","","https://careers.ltimindtree.com/","High","GIFT City office; AI-powered financial services"),
    (166,"Zensar Technologies","zensar.com","Enterprise / AI","—","careers","","","","https://www.zensar.com/careers","High","Low-code automation for healthcare & BFSI; apply via portal"),
    (167,"DevX","devx.work","Coworking / Startup Infra","—","contact","","","","","Medium","GIFT Tower 1 Gandhinagar; startup coworking; apply via website"),
    # ── ROUND 5: Mytron Labs + founder emails ────────────────────────────────
    (168,"MyTron Labs","mytronlabs.com","Physical AI / Robotics Data","founders@mytronlabs.com","contact","Priyank Patel","Co-Founder","priyank@mytronlabs.com","","High","VERIFIED ✅ — also aditya@mytronlabs.com; Physical AI data backbone"),
    # ── ROUND 6: More Gandhinagar / Infocity / Kudasan companies ─────────────
    (169,"Heptagon Global Services","heptagonservices.com","Web Dev / SEO / DevOps","sales@heptagonservices.com","contact","","","","","High","Kudasan Gandhinagar; also Canada office"),
    (170,"Hats Off Solutions","hatsoffsolutions.com","Software / Web / IoT","info@hatsoffsolutions.com","contact","","","","","High","302 Siddhraj Zori, Sargasan Gandhinagar; also USA office"),
    (171,"Quest Infosense","questinfosense.com","App Dev / Web Dev","biz@questinfosense.com","contact","","","","","High","702 Capital Icon, Sargasan Gandhinagar; also USA/Canada"),
    (172,"Dreams Technology","dreams-technology.com","Web / Mobile / Laravel","info@dreams-technology.com","contact","","","","","High","B-111 Swagat Rainforest-2, Kudasan Gandhinagar"),
    (173,"The Intech Group","theintechgroup.com","Enterprise Software / Cloud","contact@ics-global.in","contact","","","resume@ics-global.in","","High","IT Tower 3 Infocity Gandhinagar; RESUME EMAIL ✅ + Dubai office"),
    (174,"TechAvidus","techavidus.com","App Dev / Web Dev","hr@techavidus.com","careers","","","","","High","DIRECT HR EMAIL ✅ — also hello@; 405 Shikshapatri Biz Hub Kudasan"),
    (175,"eVision IT Solution","evisionits.com","CMS / Web / Digital","info@evisionits.com","contact","","","","","High","Gandhinagar + Dubai (3705 Citadel Tower Business Bay) ✅"),
    (176,"CSIT Park","csitpark.com","IT Services / Cloud / Data","hello@csitpark.com","contact","","","","","High","D-513 VTC Kudasan Gandhinagar; 25+ years"),
    (177,"Xopple Infotech","xopple.com","Web / Mobile / SEO","info@xopple.com","contact","","","","","High","C-435 Pramukh Mastana Complex Kudasan Gandhinagar"),
    (178,"Accrete Infosolution","accreteinfo.com","PHP / React / Laravel","hr@accreteinfo.com","careers","","","","","High","DIRECT HR EMAIL ✅ — also sales@; IT Tower-1 Infocity Gandhinagar"),
    (179,"Haraxy Technologies","haraxy.co","Bespoke Software / Games","info@haraxy.co","contact","","","","","High","Sargasan Gandhinagar; also letstalk@haraxy.co"),
    (180,"iPredict IT Solutions","ipredictitsolutions.com","Odoo / IT Consulting","info@ipredictitsolutions.com","contact","","","","","High","602 Pratik Mall Kudasan Gandhinagar; Odoo specialist"),
    (181,"Niharika Softweb","niharikasofttech.com","Software Dev","info@niharikasoftweb.com","contact","","","","","Medium","Pramukh Mastana Arcade Kudasan Gandhinagar"),
    (182,"AlpsLogic IT Solutions","alpslogic.in","UI/Cloud/Mobile","info@alpslogic.in","contact","","","","","Medium","415-419 Shree Ugati Corporate Park Kudasan Gandhinagar; 20+ years; .NET/SharePoint"),
    # ── ROUND 7: More Gandhinagar + GIFT City MNCs ───────────────────────────
    (183,"Computyne","computyne.com","Data/BPO/KPO","info@computyne.com","contact","","","","","High","3/2 Alpha Arcade Infocity Circle Gandhinagar; 24x7 operation"),
    (184,"Samaj Infotech","samajinfotech.com","Game Dev / Mobile / Web","info@samajinfotech.com","contact","","","","","High","VERIFIED ✅ — 435 Pramukh Mastana Complex Kudasan Gandhinagar; 100+ team; founders: Naresh & Upen Patel"),
    (185,"InfyU Labs","infyulabs.com","AgriTech / IoT / AI","hello@infyulabs.com","contact","","","","https://infyulabs.com/careers","High","VERIFIED ✅ — IIT Gandhinagar Research Park; FICCI award winner"),
    (186,"Airbow IT Services","airbow.io","Web Dev / Digital Marketing","info@airbow.io","contact","","","","","High","502 Siddhraj Z Square Kudasan Gandhinagar; also UK office"),
    (187,"Cognizant GIFT City","cognizant.com","IT Services / BFSI / AI","—","careers","","","","https://careers.cognizant.com/global-en/gift-city-interview-invitational/","High","NEW center in GIFT City; hiring Software Devs, Full Stack, Data Analysts"),
    (188,"Hexaware GIFT City","hexaware.com","TechFin / IT Services","—","careers","","","","https://hexaware.com/careers/","High","Setting up TechFin centre at GIFT City; 500 jobs announced"),
    # ── ROUND 8: Final Gandhinagar batch ─────────────────────────────────────
    (189,"C-Metric Solutions","c-metric.com","Enterprise Software / MS Partner","info@c-metric.com","contact","","","","","High","302 IT Tower-2 Infocity Gandhinagar; also sales@; USA office; Microsoft Silver Partner"),
    (190,"Bugle Technologies","bugle.in","Digital Consulting / Product Dev","sales@bugle.in","contact","","","","","High","Plot 520 Sector-1 Gandhinagar; founded 2006; also USA office"),
    (191,"Signzy","signzy.com","AI Fintech / Digital Trust","connect@signzy.com","contact","","","","https://www.signzy.com/careers","High","AI-based fintech; GIFT City presence; cross-border payments"),
    # ── ROUND 9: More Gandhinagar + Infocity + PDPU corridor ────────────────
    (192,"Shayona Infotech","shayonainfotech.com","Web / Mobile / SEO","info@shayonainfotech.com","contact","","","","","High","C-207 Business Park PDPU Road Raysan Gandhinagar; iOS/Android/CMS"),
    (193,"Kshatrainfotech","kshatrainfotech.com","ML / Image Processing / Web","hr@kshatrainfotech.com","careers","","","","","High","DIRECT HR EMAIL ✅ — 321/G Super Mall 1 Infocity Gandhinagar; ML & Image Processing focus"),
    (194,"Electrocom Technology","electrocom.in","Software Products / IoT","info@electrocom.in","contact","","","","","Medium","C-12 Electronics Zone GIDC Sector 25 Gandhinagar; software products since 2000"),
    (195,"Silver Touch Technologies","silvertouch.com","SAP / ERP / Cloud / AI","info@silvertouch.com","contact","","","","https://www.silvertouch.com/career/","High","1000+ employees; Ahmedabad HQ + Gandhinagar ops; USA/UK/Canada offices; SAP Gold Partner"),
    (196,"Dev Information Technology","devitpl.com","ERP / Cloud / Security","presales@devitpl.com","contact","","","","https://devitpl.com/careers/","High","NSE/BSE listed; CMMI Level 3; ISO 27001; 28+ years; Ahmedabad/Gandhinagar; AWS+MS partner"),
    (197,"Krify Software","krify.co","App Dev / AI / Web","hr@krify.com","careers","","","","https://krify.co/careers/","High","DIRECT HR EMAIL ✅ — Gandhinagar presence; internship program for IIIT/IIT/NIT students"),
    (198,"eInfochips (Arrow)","einfochips.com","IoT / Embedded / AI / ML","marketing@einfochips.com","contact","","","","https://careers.einfochips.com/","High","2000+ employees; Ahmedabad (near Gandhinagar); Arrow company; product engineering leader"),
    # ── ROUND 10: More Gandhinagar startups + founders ──────────────────────
    (199,"Awedus","awedus.com","HR Software / Employee Mgmt","hello@awedus.com","contact","Bhavesh Tarkhala","CEO","hello@awedus.com","","High","407 Capitol Icon Sargasan Gandhinagar; CEO: Bhavesh Tarkhala; also help@awedus.com"),
    (200,"TIS India","tisindia.com","Digital Consulting / Salesforce","hr@tisindia.com","careers","","","","https://tisindia.com/careers/","High","DIRECT HR EMAIL ✅ — Infocity Gandhinagar; web dev, digital marketing, Salesforce consulting"),
    (201,"Certopus","certopus.com","SaaS / Certificate Mgmt","support@certopus.com","contact","","","","","High","Gandhinagar-based SaaS; parent: DevSquirrel Technologies; WhatsApp: +919558817787"),
    (202,"Yudiz Solutions","yudiz.com","Game Dev / Blockchain / AR-VR / AI","contact@yudiz.com","contact","Chirag Leuva","CEO & Director","contact@yudiz.com","https://in.linkedin.com/company/yudiz-solutions-ltd","High","400+ team; Ahmedabad HQ (near Gandhinagar); USA+Canada offices; HR: +91 7874400606"),
    (203,"Cyfuture","cyfuture.com","Cloud / IT Infrastructure / BPO","info@cyfuture.com","contact","","","","https://cyfuture.com/careers.html","High","Sarkhej Gujarat office; cloud hosting, data centers, tech support BPO; 9 global locations"),
    # ── ROUND 11: Gandhinagar deep dive — sectors, founders, new finds ──────
    (204,"Lucent Innovation","lucentinnovation.com","AI / Data Engineering / Commerce","info@lucentinnovation.com","contact","Nitesh Kasma","CEO & Co-Founder","nitesh@lucentinnovation.com","https://www.linkedin.com/in/niteshkasma/","High","GIFT-SEZ Gandhinagar + USA HQ; 142 employees; Databricks Partner; CEO: Nitesh Kasma"),
    (205,"Electroware Infotech","electroware.net","FinTech / Tally / ERP","info@electroware.net","contact","","","tarangpatel@electroware.net","","High","Sector 11 Gandhinagar; since 1999; 25+ years; FinTech + Tally solutions"),
    (206,"RiYank Technologies","riyanktechnologies.com","Web / Mobile / Digital","hello@riyanktechnologies.com","contact","","","","","High","D-107 Swagat Rainforest-3 Sargasan Gandhinagar"),
    (207,"Drize Technologies","drizetech.com","Web / Cloud / Digital Marketing","info@drizetech.com","contact","","","","","Medium","Adalaj Gandhinagar; AWS/GCloud/Azure; also Thaltej office"),
    (208,"Sai Branding","saibranding.com","Digital Marketing / Branding / Web","hello@saibranding.com","contact","Nishant Darji","Founder & CEO","hello@saibranding.com","","High","509-510 Synergy Space Sargasan Gandhinagar; since 2012; Founder: Nishant Darji"),
    (209,"Realcode Infotech","realcodeinfotech.com","Web / Mobile / Bulk SMS","—","contact","Somnath Khandare","Managing Director","—","","Medium","Sector 3 Gandhinagar; since 2016; MD: Somnath Khandare"),
    # ── ROUND 12: Odoo partners + studios + deep finds ──────────────────────
    (210,"SerpentCS","serpentcs.com","Odoo ERP / CMMI3 / ISO","contact@serpentcs.com","contact","Husen Daudi & Jay Vora","Co-Founders / MDs","contact@serpentcs.com","https://www.linkedin.com/company/serpent-consulting-services","High","301 Siddhraj Zavod Sargasan Gandhinagar; Odoo top contributor; 50+ Odoo apps; since 2011"),
    (211,"Caret IT Solutions","caretit.com","Odoo ERP / Business Automation","info@caretit.com","contact","","","","https://in.linkedin.com/company/caret-it-solutions","High","706-709 Pratik Complex Kudasan Gandhinagar; Odoo GOLD Partner; India + Canada"),
    (212,"Prabhu Studio","prabhustudio.com","Animation / VFX / Web / Mobile","prabhu@prabhustudio.com","contact","Akash Pandya","Founder","prabhu@prabhustudio.com","","High","Adalaj Infocity Gandhinagar; since 2006; Animation/VFX/Web/Mobile; Founder: Akash Pandya"),
    # ── ROUND 13: More Gandhinagar + CEO emails ────────────────────────────
    (213,"Honeycomb Softwares","honeycombsoftwares.com","Web / Mobile / Cloud","contact@honeycombsoftwares.com","contact","Janak Darji","Founder & MD","janak@honeycombsoftwares.com","https://in.linkedin.com/in/janakndarji","High","A-404 Landmark Kudasan Gandhinagar; 30+ team; since 2012; Founder & CEO: Janak Darji; ZoomInfo confirmed j***@honeycombsoftwares.com"),
    (214,"Electroweb Solution","electroweb.in","Web Dev / Mobile / SEO","info@electroweb.in","contact","","","","","High","208 Shalin Centrum Sector-11 Gandhinagar; since 2012; leading IT company; Ph: +91 99040 35282"),
    (215,"Sanskar Technolab","sanskartechnolab.com","ERPNext / ERP / IT","career@sanskartechnolab.com","careers","Anand Thakker","Founder & CEO","career@sanskartechnolab.com","","High","DIRECT CAREER EMAIL ✅ — Frappe/ERPNext Partner; 64+ devs; 10+ years; also info@; HR Ph: +91 93136 55703"),
    (216,"Tiny Script Soft Tech","tinyscript.in","Web / Multimedia / IT","tinyscriptsofttech@gmail.com","contact","Akash Patel","Director","tinyscriptsofttech@gmail.com","https://in.linkedin.com/company/tiny-script-soft-tech-pvt-ltd","Medium","GIDC Vatva Ahmedabad; incorporated Sept 2022; hiring in Gandhinagar area"),
    # ── ROUND 14: Deep finds + CEO emails ──────────────────────────────────
    (217,"Rumbum Software","rumbum.co","AI / Mobile / Web / Animation","contact@rumbum.co","contact","Romit Mevada","MD & CEO","romit.mewada@gmail.com","https://in.linkedin.com/company/rumbum-software","High","B501 Shree Rang Aroma GIFT City Road Randesan Gandhinagar; since 2018; AI/Swift/Kotlin/Flutter/PHP; CEO: Romit Mevada"),
    (218,"Gurukrupa Infotech","gurukrupainfotech.com","Web / App Dev / Digital Marketing","nikul@gurukrupainfotech.com","contact","Nikul Suthar & Yuvrajsinh Vaghela","Co-Founders & CEOs","nikul@gurukrupainfotech.com","https://in.linkedin.com/in/nikul-suthar-7a24167b","Medium","211 Sharan Business Sector-26 Gandhinagar; 10-50 team; ZoomInfo: ****@gurukrupainfotech.com; Ph: 6351121958"),
    # ── ROUND 15: Odoo specialists + AI startup + GIDC finds ───────────────
    (219,"Acespritech Solutions","acespritech.com","Odoo ERP / Mobile / Web","sales@acespritech.com","contact","Navrang Oza","Founder & MD","navrang@acespritech.com","https://in.linkedin.com/company/acespritech","High","C-511 The Landmark Kudasan Gandhinagar; 13+ yrs Odoo expertise; ZoomInfo: n***@acespritech.com; Ph: +91 79902 03338"),
    (220,"Kroop AI","kroop.ai","Deepfake Detection / AI Video","kroopai@gmail.com","contact","Dr. Jyoti Joshi Dhall","CEO & Founder","kroopai@gmail.com","","High","FUNDED AI STARTUP ✅ — Gandhinagar; deepfake detection + text-to-video; Co-founders: Sarthak Gupta, Milan Chaudhari; inc42 featured"),
    (221,"Sypram Technology","sypramtechnology.com","Software Products / ACH Systems","info@sypramtechnology.com","contact","","","","","Medium","Sector 23 Gandhinagar; since 2005; also GIDC Electronic Park SEZ; COO: Dr. Manish Dholakia"),
    # ══════════════════════════════════════════════════════════════════════════
    # ██  BANGALORE / BENGALURU  ██
    # ══════════════════════════════════════════════════════════════════════════
    # ── ROUND 16: Bangalore funded startups + CEO emails ───────────────────
    (222,"SuperKalam","superkalam.com","AI EdTech / Test Prep","join@superkalam.com","careers","Vimal Singh Rathore","Founder & CEO","vimal@kalam.in","https://www.linkedin.com/in/vimal-singh-rathore/","High","YC W23 ✅ — $2M seed; AI-powered mentor for test prep; Bengaluru; also Aseem Gupta co-founder; HIRING INTERNS"),
    (223,"ClearFeed","clearfeed.ai","AI Helpdesk / Slack / Teams","hello@clearfeed.ai","contact","Joydeep Sen Sarma","Co-Founder & CEO","joydeep@clearfeed.ai","","High","$2.7M funded ✅ — Bellandur Bangalore; AI helpdesk for Slack/Teams; Co-founders: Ankit Jain, Lalit Indoria; ZoomInfo: j******@clearfeed.ai"),
    (224,"OnFinance AI","onfinance.ai","AI / Banking / BFSI Compliance","team@onfinance.in","contact","Anuj Srivastava","Co-Founder & CEO","team@onfinance.in","https://in.linkedin.com/in/anujsrivastava02","High","SEED FUNDED ✅ — Bengaluru; LLM (NeoGPT) for banking/insurance; Co-founder: Priyesh Srivastava; founded 2022"),
    (225,"Kramah Software","kramah.com","EdTech / University ERP / AI","support@kramah.com","contact","Dr. Rajeev C Raghunath","CEO & MD","rajeev.raghunath@kramah.com","https://in.linkedin.com/in/rajeevraghunath/","High","✅ DIRECT CEO EMAIL — Kumaraswamy Layout Bangalore; 85+ universities; OFFERS AI INTERNSHIPS; Ph: +91 988-005-0979"),
    (226,"Infilect","infilect.com","Computer Vision / Retail AI","careers@infilect.com","careers","Anand Prabhu Subramanian","Co-Founder & CEO","careers@infilect.com","","High","Mela Ventures funded ✅ — Koramangala Bangalore; image recognition retail analytics; OFFERS 6-MONTH INTERNSHIPS; Co-founder: Vijay Gabale"),
    (227,"WizCommerce","wizcommerce.com","B2B Commerce / AI / CRM","hello@wizcommerce.com","contact","Divyaanshu Makkar","Co-Founder & CEO","hello@wizcommerce.com","https://www.linkedin.com/in/divyaanshumakkar/","High","SERIES A ✅ — Bengaluru; B2B commerce platform for wholesale; Co-founder: Vikas Garg; founded 2020"),
    (228,"Klaar","klaarhq.com","HR Tech / Performance / SaaS","hello@klaarhq.com","contact","Sharthok Chakraborty","Co-Founder & CEO","hello@klaarhq.com","https://www.linkedin.com/in/sharthok-chakraborty","High","$6.7M Series A ✅ — WeWork Embassy TechVillage Bangalore; Agentic Performance Mgmt; Co-founder: Atri Roy; Ph: +91 983 674 0283"),
    (229,"Srishti Software","srishtisoft.com","Healthcare IT / Product","sales@srishtisoft.com","contact","Ajay Shankar Sharma","Co-Founder & CEO","sales@srishtisoft.com","","High","HSR Layout Bangalore; healthcare product PARAS; since 1997; Ph: +91 9945239357"),
    (230,"Techasoft","techasoft.com","Web / Mobile / Digital Marketing","info@techasoft.com","contact","","","","","High","HSR Layout Bangalore; 105 employees; $10-50Cr revenue; app dev + digital marketing; since 2016; Ph: +91 88847 39988"),
    # ── ROUND 17: Bangalore YC-backed + AI startups + CEO emails ─────────────
    (231,"GoodWorkLabs","goodworklabs.com","AI / ML / Mobile / Product Dev","contact@goodworklabs.com","contact","Vishwas Mudagal","Co-Founder & MD","contact@goodworklabs.com","https://in.linkedin.com/in/vishwasmudagal","High","Whitefield Bangalore; 200+ employees; AI/ML lab + outsourced product dev; CEO Sonia Sharma; angel investors; Ph: +91 80-43364621"),
    (232,"Tensorfuse","tensorfuse.io","AI Infra / Serverless GPU / MLOps","founders@tensorfuse.io","contact","Agam Jain","Co-Founder & CPO","agam@tensorfuse.io","https://www.linkedin.com/in/agam-jain-5a8b95151/","High","YC W24 ✅ — Bengaluru; serverless GPU on your own cloud; Co-founder: Samagra Sharma (CEO); samagra@tensorfuse.io"),
    (233,"GoSats","gosats.io","Fintech / Bitcoin Rewards","roshan@gosats.io","contact","Mohammed Roshan","Co-Founder & CEO","roshan@gosats.io","https://in.linkedin.com/in/roshanaslam","High","YC W22 ✅ — Bengaluru; Bitcoin stacking/rewards app; 23 employees; Co-founder: Roshni Aslam"),
    (234,"Emergent","emergent.sh","AI App Builder / No-Code","team@emergent.sh","contact","Mukund Jha","Co-Founder & CEO","team@emergent.sh","https://www.linkedin.com/in/mukundjha/","High","YC S24 ✅ — $100M Series B (Lightspeed + SoftBank); 700K+ users; $10M ARR in 2 months; ex-Dunzo CTO; twin brothers; CTO Madhav Jha (PhD CS)"),
    (235,"Flagright","flagright.com","AML Compliance / Fintech / AI","gdpr@flagright.com","contact","Madhu G Nadig","Co-Founder & CTO","gdpr@flagright.com","https://www.linkedin.com/in/madhugnadig/","High","YC W22 ✅ — $4.3M seed; AI-native AML compliance; Bangalore office; CEO Baran Ozkan; offices in NY/SF/Berlin/Singapore/Bangalore"),
    # ── ROUND 18: More Bangalore YC + AI startups ────────────────────────────
    (236,"Rivia.AI","rivia.ai","Interactive Product Demos / SaaS","samay@rivia.ai","contact","Samay Jain","Co-Founder & CEO","samay@rivia.ai","https://www.linkedin.com/in/samayjain/","High","YC S21 ✅ — Bengaluru; 5 employees; create interactive product demos in 10 min; Co-founder: Prabal Agarwal (CTO); HIRING 1 engineering role"),
    (237,"Infinity","infinityapp.in","Cross-Border Fintech / Banking","support@infinityapp.in","contact","Sourav Choraria","Co-Founder & CEO","sourav@infinityapp.in","https://www.linkedin.com/in/souravchoraria/","High","YC W24 ✅ — Bengaluru; $1.9M pre-seed; cross-border payments 70% cheaper; 15 employees; Co-founder: Sidharth Choraria; Ph: +91 95354 82864"),
    (238,"Vahan.ai","vahan.ai","AI Recruiting / WhatsApp / HR Tech","madhav@vahan.ai","contact","Madhav Krishna","Founder & CEO","madhav@vahan.ai","https://in.linkedin.com/in/madhavkrishna","High","YC ✅ — Bengaluru; $23.7M funded (Khosla + Founders Fund); AI recruiter for blue-collar; clients: Zomato/Swiggy/Uber; Columbia CS Masters; WEF Tech Pioneer"),
    (239,"smallest.ai","smallest.ai","Voice AI / TTS / STT / AI Agents","info@smallest.ai","contact","Sudarshan Kamath","Co-Founder & CEO","info@smallest.ai","https://www.linkedin.com/in/sudarshankamath/","High","Indiranagar Bangalore; voice AI models; Co-founder: Akshat Mandloi; ex-Bosch AI + Vakilsearch PM; viral Bangalore hiring post; Ph: +91 9637842074"),
    (240,"Strac","strac.io","Data Security / DLP / DSPM / AI","aatish@strac.io","contact","Aatish Mandelecha","Founder & CEO","aatish@strac.io","https://www.linkedin.com/in/aatishmandelecha/","High","YC W22 ✅ — Bengaluru office; DLP for SaaS/Cloud/GenAI; ex-Amazon 11 yrs payments infra; ACTIVELY HIRING in Bengaluru"),
    (241,"Kula","kula.ai","AI Recruitment / ATS / HR Tech","hello@kula.ai","contact","Achuthanand Ravi","Co-Founder & CEO","hello@kula.ai","https://www.linkedin.com/in/achuthanand-ravi/","High","Bengaluru + Singapore; $15M seed; AI-native ATS; ex-founding recruiter Freshworks + Uber + Stripe; Co-founders: Sathappan M, Suman Kumar Dey"),
    # ── ROUND 19: Bangalore AI/SaaS/Fintech + CEO emails ─────────────────────
    (242,"Sarvam AI","sarvam.ai","Sovereign AI / LLM / NLP","careers@sarvam.ai","careers","Pratyush Kumar","Co-Founder & CEO","careers@sarvam.ai","https://www.linkedin.com/in/pratyush-kumar-8844a8a/","High","Bengaluru; India's sovereign AI; govt contract IndiaAI Mission; Sarvam-30B/105B LLMs; Co-founder: Vivek Raghavan; ex-Google Brain; IIT Bombay PhD; 29 open roles"),
    (243,"Locale.ai","locale.ai","Geospatial Analytics / AI / SaaS","aditi@locale.ai","contact","Aditi Sinha","Co-Founder & CEO","aditi@locale.ai","https://www.linkedin.com/in/aditisinha1002/","High","Bengaluru; Forbes 30U30 ✅; $5.5M funded; $1.7M ARR; location analytics for supply chain; 11-50 employees; Co-founder: Rishabh Jain"),
    (244,"FamApp","famapp.in","Fintech / UPI / Teen Banking","careers@famapp.in","careers","Sambhav Jain","Co-Founder & CEO","sambhav@fampay.in","https://in.linkedin.com/in/sambhavanandjain","High","YC S19 ✅ — HSR Layout Bangalore; fintech for teens; Co-founder: Kush Taneja; IIT Roorkee; Forbes 30U30; HIRING creative + growth teams"),
    (245,"Rocketium","rocketium.com","AI Creative Ops / Marketing SaaS","satej@rocketium.com","contact","Satej Sirur","Co-Founder & CEO","satej@rocketium.com","https://www.linkedin.com/in/satejsirur/","High","Bangalore HQ; 78 employees; AI creative automation for enterprises; Blume Ventures funded; RETHINK Retail Top AI Leader 2024"),
    (246,"Rigi","rigi.club","Creator Economy / Monetization / SaaS","swapnil@rigi.club","contact","Swapnil Saurav","Co-Founder & CEO","swapnil@rigi.club","https://www.linkedin.com/in/linkswapnil/","High","HSR Layout Bangalore; $25M funded (Elevation Capital); 114 employees; creator community monetization; Co-founder: Ananya Singhal; prev HalaPlay (acquired)"),
    # ── ROUND 20: Bangalore DevTools / SaaS / Open Source + CEO emails ───────
    (247,"SigNoz","signoz.io","Open Source Observability / DevTools","pranay@signoz.io","contact","Pranay Prateek","Co-Founder & CEO","pranay@signoz.io","https://www.linkedin.com/in/pranay01/","High","YC W21 ✅ — Bengaluru; open source observability (OpenTelemetry); 38 employees; hiring@signoz.io; CTO Ankit Nayan"),
    (248,"Clarisights","clarisights.com","Marketing Analytics / AI / SaaS","arun@clarisights.com","contact","Arun Srinivasan","Co-Founder & CEO","arun@clarisights.com","https://www.linkedin.com/in/arun-srinivasan-clarisights/","High","✅ DIRECT CEO EMAIL — Indiranagar + Domlur Bangalore; marketing insights for Uber/HelloFresh/Delivery Hero; Co-founder: Ankur Gupta"),
    (249,"Dukaan","mydukaan.io","E-Commerce Platform / AI / SaaS","support@mydukaan.io","contact","Suumit Shah","Co-Founder & CEO","suumit@mydukaan.io","https://in.linkedin.com/in/suumitshah","High","Bengaluru; DIY e-commerce platform; famous for AI chatbot replacing support staff; Co-founder: Subhash Choudhary (CTO)"),
    (250,"Scribble Data","scribbledata.io","MLOps / Feature Store / AI","venkata@scribbledata.io","contact","Dr. Venkata Pingali","Co-Founder & CEO","venkata@scribbledata.io","https://in.linkedin.com/in/pingali","High","✅ DIRECT CEO EMAIL — Bangalore + Toronto; MLOps feature store 'Enrich'; Blume Ventures; IIT Bombay; 20 employees; HIRING"),
    (251,"Sprinto","sprinto.com","Compliance Automation / Security SaaS","sales@sprinto.com","contact","Girish Redekar","Co-Founder & CEO","girish@sprinto.com","https://www.linkedin.com/in/girishredekar/","High","Bannerghatta Road Bangalore; $32.2M Series B; autonomous compliance engine; Co-founder: Raghuveer Kancherla; prev RecruiterBox (acquired)"),
    # ── ROUND 21: Bangalore IT services + AI dev companies ───────────────────
    (252,"Krazimo","krazimo.ai","AI Engineering / Agents / Custom Dev","akhil@krazimo.ai","contact","Akhil Verghese","CEO","akhil@krazimo.ai","https://www.linkedin.com/in/akhilverghese/","High","✅ DIRECT CEO EMAIL — Bellandur/HSR Bangalore; ex-Google engineers; CTO Mridul Nagpal (ex-Google); AI agents + automation; 10-49 employees"),
    (253,"Reckonsys","reckonsys.com","AI / Chatbot / Custom Software","info@reckonsys.com","contact","Sathish Visanagiri","Founder & CEO","sathish@reckonsys.com","https://www.linkedin.com/in/sathish-visanagiri/","High","Sarjapur Road Bangalore; 50-249 employees; AI chatbots + document summarization + semantic search; Ph: +91 80613 56100"),
    (254,"Pace Wisdom","pacewisdom.com","AI Product Dev / Cloud / IoT","contact@pacewisdom.com","contact","Bharath Jatangi","Co-Founder","contact@pacewisdom.com","https://in.linkedin.com/company/pace-wisdom-solutions","High","Rajajinagar Bangalore; 50-249 employees; 150+ solutions across 10 countries; AI-first product engineering; Co-founder: Mohan Thimmadasaiah"),
    (255,"Evnek Technologies","evnek.com","Generative AI / LLM / Cloud / DevOps","info@evnek.com","contact","Ashis Kumar Sahoo","Director","info@evnek.com","https://in.linkedin.com/company/evnek","High","Whitefield (Akshay Tech Park) Bangalore; 100-200 employees; GenAI LLM solutions; deep learning + cloud; founded 2022"),
    # ── ROUND 22: Bangalore Cybersecurity + InsurTech + Benefits ──────────────
    (256,"CloudSEK","cloudsek.com","Cybersecurity / AI Threat Intel","careers@cloudsek.com","careers","Rahul Sasi","Co-Founder & CEO","rahul.sasi@cloudsek.com","https://www.linkedin.com/in/fb1h2s/","High","✅ DIRECT CEO EMAIL — Cambridge Road Bangalore; $19M Series B1; predictive cybersecurity for 250+ enterprises; 32 OPEN POSITIONS; ethical hacker turned CEO"),
    (257,"Plum","plumhq.com","InsurTech / Employee Health / SaaS","abhishek@plumhq.com","contact","Abhishek Poddar","Co-Founder & CEO","abhishek@plumhq.com","https://in.linkedin.com/in/abhishek24","High","✅ DIRECT CEO EMAIL — Bengaluru; $36M Series B; employee health benefits platform; Co-founder: Saurabh Arora; HIRING across teams"),
    (258,"Nova Benefits","novabenefits.com","InsurTech / Employee Benefits / SaaS","admin@getnovaapp.com","contact","Saransh Garg","Co-Founder & CEO","admin@getnovaapp.com","https://www.linkedin.com/in/saransh-garg/","High","Bangalore; employee wellness & insurance SaaS; Co-founder: Yash Gupta; Ph: +91 91673 39156"),
    # ── ROUND 23: Bangalore AI Voice / LegalTech / EdTech / HealthTech / Deep Tech ──
    (259,"Myelin Foundry","myelinfoundry.ai","AI Video / Edge AI / Deep Tech","social@myelinfoundry.com","contact","Gopichand Katragadda","Founder & CEO","social@myelinfoundry.com","https://in.linkedin.com/in/gkatragadda","High","Whitefield Bangalore; AI on video/voice/sensor data for edge devices; ex-Group CTO Tata Sons; ex-MD GE JFWTC; Board: Bosch Ltd, Asian Paints, ICICI Securities; Ph: +91 80 6190 4242"),
    (260,"SpotDraft","spotdraft.com","AI Contract Lifecycle / LegalTech / SaaS","shashank@spotdraft.com","contact","Shashank Bijapur","Co-Founder & CEO","shashank@spotdraft.com","https://in.linkedin.com/in/shashankbijapur","High","✅ DIRECT CEO EMAIL — HSR Layout Bangalore; $113M+ cumulative funding; $54M Series B (Vertex Growth + Prosus + Premji Invest); 234 employees; AI contract automation; Co-founders: Madhav Bhagat, Rohith Salim"),
    (261,"Murf AI","murf.ai","AI Voice / TTS / Synthetic Speech","ankur@murf.ai","contact","Ankur Edkie","Co-Founder & CEO","ankur@murf.ai","https://www.linkedin.com/in/ankuredkie/","High","✅ DIRECT CEO EMAIL — HSR Layout Bangalore; $11.5M Series A (Matrix/Z47); 6M+ users across 195 countries; IIT-KGP alumni; ex-Goldman Sachs; Co-founders: Sneha Roy (COO), Divyanshu Pandey"),
    (262,"Presentations.AI","presentations.ai","AI Presentation / Productivity / SaaS","sumanth@presentations.ai","contact","Sumanth Raghavendra","Co-Founder & CEO","sumanth@presentations.ai","https://in.linkedin.com/in/raghavendrasumanth","High","✅ DIRECT CEO EMAIL — Jayanagar Bangalore; 10M+ users; Accel-backed seed (Jan 2025); ChatGPT for presentations; 30 team members; Co-founders: Ravi Kasthuri, Saravanan Govindaraj"),
    (263,"SuperKalam","superkalam.com","AI EdTech / Test Prep / LLM","vimal@kalam.in","contact","Vimal Singh Rathore","CEO","vimal@kalam.in","https://in.linkedin.com/in/vimal-rathore","High","✅ DIRECT CEO EMAIL — Bengaluru; YC W23 ✅; $2M seed (YC + FundersClub + GoodWater); AI super-mentor for test prep; 20 employees; 11 OPEN ROLES; Co-founders: Aseem Gupta, Lakshay Nagpal; hiring: join@superkalam.com"),
    (264,"Richpanel","richpanel.com","AI Customer Support / CX / SaaS","amit@richpanel.com","contact","Amit RG","CEO & Founder","amit@richpanel.com","https://www.linkedin.com/in/amit-rg","High","Bengaluru; YC ✅ + Sequoia-backed; 2000+ brands (Ridge, Jones Road Beauty); AI agents for customer support; HIRING ML Engineer + Content roles in Bangalore"),
    (265,"Even Healthcare","even.in","HealthTech / Managed Care / AI","mayank@even.in","contact","Mayank Banerjee","Co-Founder & CEO","mayank@even.in","https://in.linkedin.com/in/mayank-banerjee-b081507b","High","✅ DIRECT CEO EMAIL — Indiranagar Bangalore; $20.8M funded ($30M Series A Khosla Ventures); subscription healthcare + own clinics/hospitals; 101-150 employees; Co-founders: Matilde Giglio, Alessandro Ialongo; Ph: +91 8047495555; also careers@even.in"),
    # ── ROUND 24: Bangalore B2B Marketing AI / Voice AI / AgriTech ────────────
    (266,"Factors.ai","factors.ai","B2B Marketing AI / Account Intelligence / SaaS","srikrishna@factors.ai","contact","Srikrishna Swaminathan","Co-Founder & CEO","srikrishna@factors.ai","https://www.linkedin.com/in/srifactorsai/","High","✅ DIRECT CEO EMAIL — BEML Layout RR Nagar Bangalore; Elevation Capital backed; ex-InMobi VP ($100M biz unit); IIM Calcutta MBA; Co-founders: Praveen Das (CPO, praveen@factors.ai), Aravind Murthy"),
    (267,"Bolna AI","bolna.ai","Voice AI / Multilingual Agents / NLP","maitreya@bolna.ai","contact","Maitreya Wagh","Co-Founder & CEO","maitreya@bolna.ai","https://www.linkedin.com/in/maitreya-wagh/","High","Bengaluru; YC F25 ✅; $6.3M seed (General Catalyst + Blume Ventures); voice AI for Indian languages; IIT Delhi + ex-Bain; Co-founder: Prateek Sachan (CTO, ex-Zomato/BrowserStack/Atlassian); open source on GitHub"),
    (268,"Fasal","fasal.co","AgriTech / AI IoT / Precision Farming","connect@wolkus.com","contact","Shailendra Tiwari","Founder & CEO","connect@wolkus.com","https://in.linkedin.com/in/shailendra-tiwari-fasal","High","Bangalore; $19.4M funded across 8 rounds; AI-powered smart irrigation + horticulture platform; 107 employees; 12,000+ farmers across India & SE Asia; Co-founder: Ananda Prakash Verma"),
    (269,"Skit.ai","skit.ai","Voice AI / Conversational AI / Debt Collection","scale@skit.ai","careers","Sourabh Gupta","Co-Founder & CEO","scale@skit.ai","https://www.linkedin.com/in/sourabhsg/","High","Old Madras Road Bangalore + NYC HQ; Forbes 30U30 Asia ✅; voice AI for contact centers; hiring Founder Office role in Bengaluru; Co-founder: Akshay Deshraj (CTO); IIT Roorkee"),
    # ── ROUND 25: Bangalore AI CX / Inference Infra / Marketing Cloud ─────────
    (270,"Hiver","hiverhq.com","AI Customer Service / Email Helpdesk / SaaS","niraj@hiverhq.com","contact","Niraj Ranjan Rout","Founder & CEO","niraj@hiverhq.com","https://www.linkedin.com/in/nirajranjan/","High","HSR Layout Bangalore; $46.2M funded (K1 Investment + Kalaari Capital); AI helpdesk built for Google Workspace; 10,000+ teams globally; IIT KGP; Co-founder: Nitesh Nandy; prev co-founded Mobicules"),
    (271,"Simplismart","simplismart.ai","AI Inference Infra / MLOps / GPU Cloud","amritanshu@simplismart.ai","contact","Amritanshu Jain","Co-Founder & CEO","amritanshu@simplismart.ai","https://www.linkedin.com/in/jainamritanshu/","High","Richmond Town Bangalore; $14M funded (Accel + Titan Capital); fastest AI inference engine; ex-Oracle ML Engineer; BITS Pilani; Co-founder: Devansh Ghatak (CTO)"),
    (272,"Pixis","pixis.ai","AI Marketing Cloud / Codeless AI Infra","shubham@pixis.ai","contact","Shubham A Mishra","Co-Founder & Global CEO","shubham@pixis.ai","https://www.linkedin.com/in/shubhammishra01/","High","Bengaluru + Burlingame; AI-powered marketing infrastructure; 201-500 employees; BITS Pilani; prev co-founded Absentia VR; Co-founders: Vrushali Prasade (CTO), Harikrishna Valiyath"),
    # ── ROUND 26: Bangalore Digital Adoption / Open Source DevTools ───────────
    (273,"Whatfix","whatfix.com","Digital Adoption Platform / AI / SaaS","khadim@whatfix.com","contact","Khadim Batti","Co-Founder & CEO","khadim@whatfix.com","https://in.linkedin.com/in/khadim","High","Bengaluru HQ; 800+ employees; 700+ enterprises incl Fortune 500; digital adoption + no-code analytics; CEO based in Bengaluru; IIIT-B governing body; Co-founder: Vara Kumar; ZoomInfo confirmed k***@whatfix.com"),
    (274,"ToolJet","tooljet.com","Open Source Low-Code / AI Apps / DevTools","navaneeth@tooljet.com","contact","Navaneeth Padanna Kalathil","Founder & CEO","navaneeth@tooljet.com","https://www.linkedin.com/in/navaneeth-pk/","High","✅ DIRECT CEO EMAIL — Bangalore-rooted open source low-code platform; GitHub 35K+ stars; ACTIVELY HIRING engineers + PMs + designers; Ph: +91 9400812423; also navaneethpk@outlook.com"),
    (275,"Hasura (PromptQL)","hasura.io","GraphQL API / AI Data Access / DevTools","tanmai@hasura.io","contact","Tanmai Gopal","Co-Founder & CEO","tanmai@hasura.io","https://www.linkedin.com/in/tanmaig/","High","Bangalore + SF offices; GraphQL engine + PromptQL AI; IIT Madras; $136M+ funded (unicorn 2022); Co-founder: Rajoshi Ghosh; HIRING via hasura.io/careers"),
    # ── ROUND 27: Bangalore SpaceTech + Gaming Studios ────────────────────────
    (276,"Pixxel","pixxel.space","SpaceTech / Hyperspectral Satellites / AI","awais@pixxel.space","contact","Awais Ahmed","Founder & CEO","awais@pixxel.space","https://in.linkedin.com/in/awaisahmedna","High","Bengaluru; $96M funded over 11 rounds; world's highest-res hyperspectral satellite constellation (Firefly); 177 employees; BITS Pilani; Co-founder: Kshitij Khandelwal (CTO)"),
    (277,"GalaxEye","galaxeye.space","SpaceTech / Multi-Sensor Imaging / SAR","contact@galaxeye.space","contact","Suyash Singh","Founder & CEO","suyash@galaxeye.space","https://in.linkedin.com/in/suyashaoc","High","🎯 SEEKING INTERNS — Bengaluru; Drishti Mission all-weather imaging satellite; IIT Madras spin-off; co-founders: Denil Chawda, Kishan Thakkar, Pranit Mehta, Rakshit Bhatt; careers.galaxeye.space"),
    (278,"Digantara","digantara.co.in","SpaceTech / Space Surveillance / Defence","anirudh@digantara.co.in","contact","Anirudh Sharma","Co-Founder & CEO","anirudh@digantara.co.in","https://in.linkedin.com/in/anirudh-sharma-digantara","High","✅ DIRECT CEO EMAIL — Hebbal Bangalore; $50M raised Dec 2025 (space-based missile defence); world's first commercial space surveillance satellite (SCOT); HIRING engineers + design + marketing; Co-founder: Rahul Rawat"),
    (279,"Bombay Play","bombayplay.com","Gaming / Casual Games / Studio","sruthi@bombayplay.com","careers","Oliver Jones","Co-Founder & CEO","sruthi@bombayplay.com","https://www.linkedin.com/in/oliverjones-games/","High","✅ DIRECT HIRING EMAIL — Indiranagar Bangalore; casual game studio; send resume to sruthi@bombayplay.com; Co-founder: Abhas Saroha"),
    (280,"Mayhem Studios","mayhem-studios.com","Gaming / AAA Mobile / Battle Royale","careers@mayhem-studios.com","careers","Ojas Vipat","Founder & CEO","careers@mayhem-studios.com","https://www.linkedin.com/in/ojasvipat/","High","Bangalore; India's first AAA mobile gaming studio; backed by MPL (Mobile Premier League); built Underworld Gang Wars (UGW)"),
    # ── ROUND 28: Bangalore EV / Mobility startups ────────────────────────────
    (281,"Ultraviolette Automotive","ultraviolette.com","EV / Electric Motorcycles / Embedded AI","contact@ultraviolette.com","contact","Narayan Subramaniam","Co-Founder & CEO","narayan@ultraviolette.com","https://in.linkedin.com/in/narayan-s","High","Domlur Bangalore; F77 electric motorcycle maker; TVS-backed; competing with KTM/BMW; embedded software + battery AI roles; Co-founder: Niraj Rajmohan (CTO)"),
    (282,"River Mobility","rideriver.com","EV / Electric Scooters / Mobility","aravind@rideriver.com","contact","Aravind Mani","Co-Founder & CEO","aravind@rideriver.com","https://in.linkedin.com/in/aravindmani","High","Bengaluru; Indie electric scooter; backed by Yamaha + Maniv Mobility + TrucksVC; ZoomInfo confirmed a***@rideriver.com; Co-founder: Vipin George (CTO)"),
    (283,"Simple Energy","simpleenergy.in","EV / Electric Scooters / Battery Tech","info@simpleenergy.in","contact","Suhas Rajkumar","Founder & CEO","suhas@simpleenergy.in","https://in.linkedin.com/in/suhas-rajkumar-277824150","High","Yelahanka Bangalore; Simple One scooter (240km range); raised ₹250Cr to scale production; targeting 10x sales growth; software + BMS + connected vehicle roles"),
    # ── ROUND 29: Bangalore Emotion AI / Conversational AI / Sovereign LLM ────
    (284,"Entropik","entropik.io","Emotion AI / Consumer Insights / SaaS","info@entropiktech.com","contact","Ranjan Kumar","Co-Founder & CEO","ranjan@entropiktech.com","https://www.linkedin.com/in/ranjan-kr/","High","Bengaluru; Series B; global leader in Emotion AI; unified human insights platform; ZoomInfo confirmed R***@entropiktech.com; Ph: 080-4375-9863; Co-founders: Lava Kumar, Bharat Singh Shekhawat"),
    (285,"Senseforth.ai","senseforth.ai","Conversational AI / NLP / Chatbots","contact@senseforth.ai","contact","Shridhar Marri","Co-Founder & CEO","shridhar@senseforth.ai","https://in.linkedin.com/in/shridharmarri","High","Girinagar Bangalore; enterprise conversational AI (banks, insurance); Co-founders: Krishna Kadiri, Ritesh Radhakrishnan (CTO); email pattern estimated"),
    (286,"CoRover.ai","corover.ai","Conversational AI / BharatGPT / Sovereign LLM","contact@corover.ai","contact","Ankush Sabharwal","Founder & CEO","ankush@corover.ai","https://in.linkedin.com/in/ankushsabharwal","High","Bangalore; built BharatGPT (India's first LLM GenAI); 1B+ users served (IRCTC bot); 50,000+ enterprises; Series A (Canbank VC); Co-founders: Manav Gandotra, Kunal Bhakhri; email pattern estimated"),
    # ── ROUND 30: Bangalore HR Tech / Developer Hiring / CX Analytics ─────────
    (287,"Springworks","springworks.in","HR Tech / Background Verification / SaaS","kartik@springworks.in","contact","Kartik Mandaville","Founder & CEO","kartik@springworks.in","https://www.linkedin.com/in/kartik-mandaville-springworks/","High","Bengaluru + Santa Monica; remote-first ALWAYS HIRING (springworks.springrecruit.com); products: SpringVerify, SpringRecruit, EngageWith, Trivia; CMU grad; email pattern estimated"),
    (288,"HackerEarth","hackerearth.com","Developer Assessment / Hiring Tech / SaaS","support@hackerearth.com","contact","Vikas Aditya","CEO","support@hackerearth.com","https://www.linkedin.com/in/vikasaditya/","High","Bengaluru; developer skill assessment + coding interviews; founded by IIT Roorkee alumni (Sachin Gupta, Vivek Prakash); 7M+ developer community"),
    (289,"Clootrack","clootrack.com","CX Analytics / AI Insights / SaaS","shameel@clootrack.com","contact","Shameel Abdulla","Co-Founder & CEO","shameel@clootrack.com","https://www.linkedin.com/in/shameelabdulla/","High","Bangalore; $4M raised (Inventus + Salesforce Ventures); AI customer experience analytics; serial entrepreneur; Co-founder: Subbakrishna Rao (CTO); email pattern estimated"),
    # ── ROUND 31: Bangalore Sales AI / WealthTech ─────────────────────────────
    (290,"Salesken","salesken.ai","Sales AI / Conversational Intelligence / SaaS","surga@salesken.ai","contact","Surga Thilakan","Co-Founder & CEO","surga@salesken.ai","https://in.linkedin.com/in/surga-thilakan-0196994","High","✅ DIRECT CEO EMAIL — Bengaluru; Series B; $41M raised over 11 rounds (Microsoft M12 + Sequoia); in-call live sales intelligence; ex-Goldman Sachs; IIM-A MBA; Co-founder: Sreeraman Vaidyanathan; Ph: +91 80416 49503"),
    (291,"Smallcase","smallcase.com","WealthTech / Investment Platform / Fintech","vasanth@smallcase.com","contact","Vasanth Kamath","Founder & CEO","vasanth@smallcase.com","https://in.linkedin.com/in/vasanthskamath","High","Richmond Road Bangalore; 367 employees; changing how India invests; IIT-KGP founders; Co-founders: Anugrah Shrivastava, Rohan Gupta; Blume + Sequoia backed; email pattern estimated"),
    # ── ROUND 32: Gandhinagar second pass — web tech with HR numbers ──────────
    (292,"BOSC Tech Labs","bosctechlabs.com","Mobile / Web App Development","hr@bosctechlabs.com","careers","HR Team","HR","hr@bosctechlabs.com","https://in.linkedin.com/company/bosc-tech-labs","High","✅ DIRECT HR EMAIL — Sargasan Gandhinagar; 5 OPEN JOBS on Indeed; Flutter/React specialists; also info@ + contact@bosctechlabs.com; founded 2017"),
    (293,"Samcom Technologies","samcomtechnologies.com","AI Voice / Chat / Workflow Automation","ketan@samcomtechnologies.com","contact","Ketan Jadhav","Founder & CEO","ketan@samcomtechnologies.com","https://theorg.com/org/samcom-technologies/org-chart/ketan-jadhav","High","Motera Ahmedabad (Gandhinagar border); AI voice/chat/workflow automation; 20+ yrs experience; founded 2013; ZoomInfo confirmed k***@samcomtechnologies.com"),
    (294,"Evolvision Technologies","evolvision.com","Web / Software Development","business@evolvision.com","contact","Jimit Joshi","Founder & CEO","business@evolvision.com","https://in.linkedin.com/in/jimitjoshi","High","✅ PHONE: +91-9510645454 — Pramukh Arcade-2 Kudasan Gandhinagar; partners: Jimit Joshi + Sandip Patel"),
    (295,"Unity Infoway","unityinfoway.com","Web / Mobile Development","info@unityinfoway.com","contact","Rahul Gondaliya","Founder","info@unityinfoway.com","https://in.linkedin.com/company/unity-infoway","High","Radhe Infinity, Rakshashakti Circle, Kudasan Gandhinagar; serving since 2011; email estimated"),
]

hr_contacts = [
    ("Bayut","Mirna Al Sayegh","Sr. Talent Acquisition Specialist","m***@bayut.com","https://ae.linkedin.com/in/mirna-al-sayegh-assoc-cipd-72baaa42","DM on LinkedIn + email careers@bayut.com. Use Apollo.io to reveal full email."),
    ("Huspy","John Michael Razon","People Team","john.razon@huspy.io","https://ae.linkedin.com/in/john-michael-razon-ba249066","VERIFIED - email directly"),
    ("Huspy","Grasielly D.","People Lead","grasielly@huspy.io (estimated)","https://ae.linkedin.com/in/grasielly-d-06136292","DM on LinkedIn to confirm email"),
    ("CAFU","Pranita Talukdar","HR Manager","p***@cafu.com","","DM on LinkedIn + email hello@cafu.com. Use Apollo.io to reveal full email."),
    ("Alaan","—","—","careers@alaanpay.com","https://www.alaan.com/careers","VERIFIED direct careers email - send resume here even with no open role"),
    ("Enova by Veolia","—","Recruitment Team","recruitment@enova-me.com","https://www.linkedin.com/posts/enova-me_hiring-uae-cv-activity-7449367881340940288-ZiRc","VERIFIED from LinkedIn post — active hiring email confirmed April 2026"),
    ("Melodica UAE","Sasha W.","Hiring Manager","sasha.w@melodica.ae","","VERIFIED from LinkedIn post — named contact for Data Analyst hiring"),
    ("Tabby","Anastasija Konoreva","Talent Acquisition Partner","a.konoreva@tabby.ai (use Apollo)","https://ae.linkedin.com/in/konoreva","ZoomInfo confirmed @tabby.ai — use Apollo to reveal full email"),
    ("Tabby","Ohoud Anwar","TA Associate","o.anwar@tabby.ai (use Apollo)","","ZoomInfo confirmed @tabby.ai — DM on LinkedIn"),
    ("Tabby","Mariam Mohsen","TA Coordinator","m.mohsen@tabby.ai (use Apollo)","","ZoomInfo confirmed @tabby.ai — DM on LinkedIn"),
    ("Tabby","Stephen Collopy","Talent Acquisition Lead","s.collopy@tabby.ai (use Apollo)","https://www.linkedin.com/in/stephencollopysc/","TA Lead at Tabby Dubai — DM + Apollo"),
    ("Ziina","Anton Taranenko","Global TA Lead","a.taranenko@ziina.com (use Apollo)","https://www.linkedin.com/in/antontaranenko/","TA Lead at Ziina Dubai — DM on LinkedIn"),
    ("Ziina","Sara Latorre","People Team","s.latorre@ziina.com (use Apollo)","https://ae.linkedin.com/in/sara-latorre-6b147a86","People Team at Ziina — DM on LinkedIn"),
    ("Qashio","Ashley McBean","Talent Acquisition Lead","a.mcbean@qashio.com (use Apollo)","https://www.linkedin.com/in/ashleymcbean/","TA Lead at Qashio Dubai — DM + Apollo"),
    ("Bayut / Dubizzle","Omar Khouly","TA Specialist","omar.khouly@bayut.com","https://www.linkedin.com/in/omar-elkhouly","✅ VERIFIED from LinkedIn post — email appeared in hiring post text directly"),
    ("Bayut / Dubizzle","Kazi Asfahan","HR | TA Team","kazi.asfahan@bayut.com (use Apollo)","https://www.linkedin.com/in/kazi-asfahan-735545248","Found via Mirna's reposts — scrape Apollo for full email"),
    ("SQUATWOLF","Ann Maria Thekkanath Johnson","People & Culture Trainee","ann.maria@squatwolf.com (use Apollo)","https://www.linkedin.com/in/ann-maria-thekkanath-johnson-7b4ab0247","Found in Bayut TA suggested contacts — DM on LinkedIn"),
    ("Adapts Media","Ankita","Manager","ankita@adaptsmedia.com","","✅ VERIFIED — scraped directly from contact page"),
    ("Narola Infotech","Raj","—","raj@narolainfotech.com","","✅ VERIFIED — scraped directly from contact page"),
    # ── CAREEM TA Team (scraped from LinkedIn company people page) ────────────
    ("Careem","Zara Iqbal","Talent Acquisition Partner","zara.iqbal@careem.com","https://www.linkedin.com/in/zara-iqbal/","Pattern: firstname.lastname@careem.com — DM + Apollo to verify"),
    ("Careem","Sanya Imran","Talent Acquisition - Tech","sanya.imran@careem.com","https://www.linkedin.com/in/sanya-imran/","Tech hiring focus — ideal contact"),
    ("Careem","Ranwa Aboulaben","Regional Senior TA Manager","ranwa.aboulaben@careem.com","https://www.linkedin.com/in/ranwa-aboulaben/","Senior manager — best for internship ask"),
    ("Careem","Javeria Taufique","Talent Acquisition","javeria.taufique@careem.com","https://www.linkedin.com/in/javeria-taufique/","Pattern: firstname.lastname@careem.com"),
    ("Careem","Zoha Binte Akmal Shariff","Talent Acquisition","zoha.shariff@careem.com","https://www.linkedin.com/in/zoha-binte-akmal-shariff/","Pattern: firstname.lastname@careem.com"),
    # ── G42 TA Team ───────────────────────────────────────────────────────────
    ("G42","Tabbashum Khan","Talent Acquisition Specialist","tabbashum.khan@g42.ai","https://www.linkedin.com/in/tabbashum-khan/","AI company Abu Dhabi — use Apollo to verify"),
    ("G42","Paul Beard","Talent Acquisition","paul.beard@g42.ai","https://www.linkedin.com/in/paul-beard-g42/","Pattern: firstname.lastname@g42.ai"),
    ("G42","Sagar Tanwar","Talent Acquisition","sagar.tanwar@g42.ai","https://www.linkedin.com/in/sagar-tanwar-g42/","'Hiring Talents at G42' — direct contact"),
    ("G42","Giselle Melissa Correa","TA / HR Operations","giselle.correa@g42.ai","https://www.linkedin.com/in/giselle-melissa-correa/","HR Ops at G42 — DM on LinkedIn"),
    # ── Property Finder TA Team ───────────────────────────────────────────────
    ("Property Finder","Neelam Anwar","Senior TA Partner - Tech & Product","neelam.anwar@propertyfinder.ae","https://www.linkedin.com/in/neelam-anwar-pf/","Tech & Product hiring — best match for you"),
    ("Property Finder","Soukaina Khassal","Senior TA Leader","soukaina.khassal@propertyfinder.ae","https://www.linkedin.com/in/soukaina-khassal/","Senior TA — email + DM"),
    ("Property Finder","Shilpa Reddy Lakka","VP Talent Acquisition","shilpa.lakka@propertyfinder.ae","https://www.linkedin.com/in/shilpa-reddy-lakka/","VP level — highest authority, email directly"),
    ("Property Finder","Ben Wilson","Senior Executive Recruitment Partner","ben.wilson@propertyfinder.ae","https://www.linkedin.com/in/ben-wilson-pf/","Executive recruiter — DM on LinkedIn"),
    # ── Kitopi TA ─────────────────────────────────────────────────────────────
    ("Kitopi","Nouha Fakhoury","Talent Acquisition Manager","nouha.fakhoury@kitopi.com","https://www.linkedin.com/in/nouha-fakhoury/","TA Manager Dubai — email + DM"),
    # ── Noon TA ───────────────────────────────────────────────────────────────
    ("Noon","Hima Gireesh","HR Coordinator & TA Specialist","hima.gireesh@noon.com","https://www.linkedin.com/in/hima-gireesh/","Dubai based — DM + Apollo"),
    ("Noon","Rafshana Mohammed Ali","TA & HR Generalist","rafshana.ali@noon.com","https://www.linkedin.com/in/rafshana-mohammed-ali/","100+ hires/year — very active recruiter"),
    # ── Talabat TA ────────────────────────────────────────────────────────────
    ("Talabat","Yasmin Atef Samir","Sr. Data Recruiter — AI & ML Hiring","yasmin.samir@talabat.com","https://www.linkedin.com/in/yasmin-atef-samir/","Hires AI/ML Engineers — perfect match"),
    ("Talabat","Mohanad El Mughrabi","Talent Acquisition Partner","mohanad.elmughrabi@talabat.com","https://www.linkedin.com/in/mohanad-el-mughrabi/","Dubai based — DM + Apollo"),
    # ── Bayzat TA ─────────────────────────────────────────────────────────────
    ("Bayzat","Manal Saleh","Talent Acquisition Coordinator","manal.saleh@bayzat.com","https://www.linkedin.com/in/manal-saleh-bayzat/","Dubai based TA Coordinator"),
    ("Bayzat","Sanjo Joshi","Sr. Talent Resourcer","sanjo.joshi@bayzat.com (estimated)","","DM on LinkedIn to confirm email"),
    # ── Gandhinagar / GIFT City Founders & HR ────────────────────────────────
    ("MyTron Labs","Priyank Patel","Co-Founder","priyank@mytronlabs.com","","✅ VERIFIED — scraped from contact page; Physical AI startup"),
    ("MyTron Labs","Aditya Gupta","Co-Founder","aditya@mytronlabs.com","","✅ VERIFIED — scraped from contact page; also founders@mytronlabs.com"),
    ("DRC Systems","Careers / HR Team","HR Department","careers@drcsystems.com","","✅ VERIFIED — scraped from contact page; also sales@, ir@, press@drcsystems.com"),
    ("DRC Systems","UAE Office","Dubai Branch","info@drcsystems.ae","","✅ VERIFIED — Dubai office: The Meydan, Nad Al Sheba Rd"),
    ("Argusoft India","Jobs / HR","Recruitment","jobs@argusoft.com","https://careers.argusoft.com","✅ VERIFIED — dedicated jobs email; A66 GIDC Gandhinagar"),
    ("Bosc Tech Labs","HR Department","HR","hr@bosctechlabs.com","","✅ VERIFIED — Kudasan Gandhinagar; also contact@bosctechlabs.com"),
    ("Cybage Software","Careers Team","Careers","careers@cybage.com","https://www.cybage.com/careers","✅ VERIFIED — Tower II Infocity Gandhinagar"),
    ("Anblicks","Careers Team","Careers","careers@anblicks.com","","✅ VERIFIED — also info@ and marketing@anblicks.com"),
    ("AvenuesAI (Infibeam)","Contact Team","General","contactus@avenuesai.com","","✅ VERIFIED — also ir@, corpcom@avenuesai.com; GIFT City"),
    # ── Round 9: More Gandhinagar HR contacts ────────────────────────────────
    ("Kshatrainfotech","HR Team","HR","hr@kshatrainfotech.com","","✅ VERIFIED — Infocity Gandhinagar; ML/Image Processing focus; also support@kshatrainfotech.com"),
    ("Krify Software","HR Team","HR / Internships","hr@krify.com","https://krify.co/careers/","✅ VERIFIED — internship program for IIIT/IIT/NIT students; Python, Mobile, Data Analytics"),
    ("Silver Touch Technologies","CS Team","Client Services","info@silvertouch.com","https://www.silvertouch.com/career/","SAP Gold Partner; 1000+ employees; Gandhinagar + global offices"),
    ("AlpsLogic IT Solutions","Contact Team","General","info@alpslogic.in","","VERIFIED — 20+ years; .NET/SharePoint/Mobile; Kudasan Gandhinagar"),
    ("Shayona Infotech","Contact Team","General","info@shayonainfotech.com","","VERIFIED — PDPU Road Raysan Gandhinagar; Founder: Devdattsinh Raol"),
    # ── Round 10: Founders & HR contacts ─────────────────────────────────────
    ("Awedus","Bhavesh Tarkhala","CEO","hello@awedus.com","","CEO of Awedus; HR/Employee mgmt SaaS; 407 Capitol Icon Sargasan Gandhinagar"),
    ("TIS India","HR Team","HR","hr@tisindia.com","https://tisindia.com/careers/","✅ VERIFIED — Infocity Gandhinagar; Digital consulting + Salesforce"),
    ("Samaj Infotech","Naresh Patel & Upen Patel","Co-Founders","info@samajinfotech.com","","✅ VERIFIED — 100+ team game dev studio; Kudasan Gandhinagar"),
    ("Yudiz Solutions","Chirag Leuva","CEO & Director","contact@yudiz.com","https://in.linkedin.com/company/yudiz-solutions-ltd","400+ team; Game/Blockchain/AR-VR; HR phone: +91 7874400606"),
    ("Certopus","Vraj (Team Lead)","Demos/Sales","support@certopus.com","","Gandhinagar SaaS startup; certificate management platform; WhatsApp: +919558817787"),
    # ── Round 11: More founders ──────────────────────────────────────────────
    ("Lucent Innovation","Nitesh Kasma","CEO & Co-Founder","nitesh@lucentinnovation.com","https://www.linkedin.com/in/niteshkasma/","GIFT-SEZ Gandhinagar; 142 employees; Databricks Partner; phone: +91 9426212181"),
    ("Electroware Infotech","Tarang Patel","Director (likely)","tarangpatel@electroware.net","","Sector 11 Gandhinagar; FinTech/Tally specialist since 1999"),
    ("Sai Branding","Nishant Darji","Founder & CEO","hello@saibranding.com","","Sargasan Gandhinagar; digital marketing + branding since 2012"),
    # ── Round 12: Odoo partners + studio founders ────────────────────────────
    ("SerpentCS","Husen Daudi","Co-Founder & MD","contact@serpentcs.com","https://www.linkedin.com/company/serpent-consulting-services","Sargasan Gandhinagar; Odoo top contributor since 2011; CMMI3+ISO certified"),
    ("SerpentCS","Jay Vora","Co-Founder & MD","contact@serpentcs.com","","Co-founded with Husen Daudi; working on Odoo since 2007 (TinyERP era)"),
    ("Prabhu Studio","Akash Pandya","Founder","prabhu@prabhustudio.com","","Adalaj/Infocity Gandhinagar; Animation/VFX/Web since 2006; phone: +91 9824264610"),
    # ── Round 13: More Gandhinagar founders + CEO emails ─────────────────
    ("Honeycomb Softwares","Janak Darji","Founder & MD","janak@honeycombsoftwares.com","https://in.linkedin.com/in/janakndarji","Kudasan Gandhinagar; 30+ team; 15+ yrs IT exp; ZoomInfo: j***@honeycombsoftwares.com; also contact@honeycombsoftwares.com"),
    ("Sanskar Technolab","Anand Thakker","Founder & CEO","career@sanskartechnolab.com","","ERPNext specialist; 64+ devs; 10+ yrs exp; also info@sanskartechnolab.com; HR: +91 93136 55703"),
    ("DRC Systems","Hiten Barchha","CEO & Managing Director","hiten@drcsystems.com","","✅ DIRECT CEO EMAIL — GIFT City 24th Floor; 300+ employees; public company (NSE listed)"),
    ("DRC Systems","Kirit Gajera","Co-Founder & COO","","","Co-Founder of DRC Systems; also COO; GIFT City Gandhinagar"),
    # ── Round 14: Batch founder contacts for existing Gandhinagar companies ──
    ("Rumbum Software","Romit Mevada","MD & CEO","romit.mewada@gmail.com","","B501 Shree Rang Aroma GIFT City Rd Randesan Gandhinagar; also contact@rumbum.co; Ph: +91 9106364543"),
    ("Gurukrupa Infotech","Nikul Suthar","Founder & CEO","nikul@gurukrupainfotech.com","https://in.linkedin.com/in/nikul-suthar-7a24167b","Sector 26 Gandhinagar; ZoomInfo: ****@gurukrupainfotech.com"),
    ("Gurukrupa Infotech","Yuvrajsinh Vaghela","Founder & CEO","yuvrajsinh@gurukrupainfotech.com","https://in.linkedin.com/in/yuvrajsinh-vaghela","ZoomInfo: v***@gurukrupainfotech.com; Ph: 6351121958"),
    ("Dreams Technology","Chetan Makwana","Founder","chetan@dreams-technology.com","https://in.linkedin.com/in/chetan-makwana-99480339","ZoomInfo: c***@dreams-technology.com; B-111 Swagat Rainforest-2 Kudasan Gandhinagar"),
    ("Oddeven Infotech","Tejpal Navadiya","Founder & CEO","hello@oddeveninfotech.com","https://in.linkedin.com/in/tejpalnavadiya","Infocity Gandhinagar; Salesforce/AI/digital transformation specialist"),
    ("iPredict IT Solutions","Ajay Patel","Co-Founder & CEO","hr@ipredictitsolutions.com","","602 Pratik Mall Kudasan Gandhinagar; Odoo/SugarCRM specialist; Ph: +91 8401557997"),
    ("iPredict IT Solutions","Bhavik Bagadiya","Co-Founder & CEO","hr@ipredictitsolutions.com","","Co-Founded with Ajay Patel; Odoo ERP specialist"),
    ("Niharika Softweb","Vijaykumar Gadhavi","Director & Founder","vijaygadhavi65@gmail.com","","Pramukh Mastana Arcade Kudasan Gandhinagar; 18 employees; also info@niharikasoftweb.com"),
    ("Quest Infosense","Rushi Gothaliya","CEO","biz@questinfosense.com","","702 Capital Icon Sargasan Gandhinagar; also USA/Canada; $6M yearly revenue"),
    # ── Round 15: More founder contacts ──────────────────────────────────────
    ("Acespritech Solutions","Navrang Oza","Founder & MD","navrang@acespritech.com","","Kudasan Gandhinagar; 13+ yrs Odoo; ZoomInfo: n***@acespritech.com; also sales@acespritech.com"),
    ("Kroop AI","Dr. Jyoti Joshi Dhall","CEO & Founder","kroopai@gmail.com","","FUNDED Gandhinagar AI startup; deepfake detection; featured on inc42 & indiaai.gov.in"),
    ("Kroop AI","Sarthak Gupta","Co-Founder","kroopai@gmail.com","","Co-founded Kroop AI with Jyoti Joshi & Milan Chaudhari"),
    # ── Round 15b: CEO emails for existing companies ─────────────────────────
    ("Hats Off Solutions","Mayank Parmar","CEO","mayank@hatsoffsolutions.com","","✅ DIRECT CEO EMAIL — 302 Siddhraj Zori Sargasan Gandhinagar; also info@hatsoffsolutions.com"),
    ("Prismetric","Ashish Parmar","Co-Founder & CEO","biz@prismetric.com","","SF/215 IT Tower 1 Infocity Gandhinagar; Ph: +91-79-40070367; Co-founded with Ketan Chavda"),
    ("Prismetric","Ketan Chavda","Co-Founder & MD","biz@prismetric.com","","Infocity Gandhinagar; Managing Director; Co-founded with Ashish Parmar"),
    ("Bosc Tech Labs","Mahesh Lalwani","Founder & CEO","mahesh@bosc.in","https://in.linkedin.com/in/mahesh-lalwani-43803a25","✅ DIRECT CEO EMAIL — Sargasan Gandhinagar; also contact@bosc.in; est 2019"),
    # ══════ BANGALORE FOUNDER/CEO CONTACTS ══════════════════════════════════
    ("SuperKalam","Vimal Singh Rathore","Founder & CEO","vimal@kalam.in","","✅ DIRECT CEO EMAIL — YC W23; $2M seed; prev founded Coursavy (acquired by Unacademy); also join@superkalam.com"),
    ("ClearFeed","Joydeep Sen Sarma","Co-Founder & CEO","joydeep@clearfeed.ai","","$2.7M funded; Bellandur Bangalore; ZoomInfo: j******@clearfeed.ai; AI helpdesk Slack/Teams"),
    ("OnFinance AI","Anuj Srivastava","Co-Founder & CEO","team@onfinance.in","https://in.linkedin.com/in/anujsrivastava02","Seed funded; NeoGPT for banking; also co-founder Priyesh Srivastava"),
    ("Kramah Software","Dr. Rajeev C Raghunath","CEO & MD","rajeev.raghunath@kramah.com","https://in.linkedin.com/in/rajeevraghunath/","✅ DIRECT CEO EMAIL — 22+ yrs exp; ex-Oracle/ITC InfoTech; 85+ universities; Ph: +91 988-005-0979"),
    ("Infilect","Anand Prabhu Subramanian","Co-Founder & CEO","careers@infilect.com","","Koramangala Bangalore; patented CV tech; Mela Ventures backed; OFFERS 6-MONTH INTERNSHIPS"),
    ("WizCommerce","Divyaanshu Makkar","Co-Founder & CEO","hello@wizcommerce.com","https://www.linkedin.com/in/divyaanshumakkar/","Series A; B2B commerce for wholesale distributors"),
    ("Klaar","Sharthok Chakraborty","Co-Founder & CEO","hello@klaarhq.com","https://www.linkedin.com/in/sharthok-chakraborty","$6.7M Series A; Agentic Performance Management; WeWork Embassy TechVillage Bangalore"),
    ("Srishti Software","Ajay Shankar Sharma","Co-Founder & CEO","sales@srishtisoft.com","","HSR Layout Bangalore; healthcare product PARAS since 1997; Ph: +91 9945239357"),
    # ── Round 17: More Bangalore CEO contacts ───────────────────────────────
    ("GoodWorkLabs","Vishwas Mudagal","Co-Founder & MD","contact@goodworklabs.com","https://in.linkedin.com/in/vishwasmudagal","Whitefield Bangalore; serial entrepreneur + bestselling author; angel investor; also CEO Sonia Sharma"),
    ("GoodWorkLabs","Sonia Sharma","Co-Founder & CEO","contact@goodworklabs.com","","Co-founded GoodWorkLabs; technopreneurs; AI/ML lab"),
    ("Tensorfuse","Agam Jain","Co-Founder & CPO","agam@tensorfuse.io","https://www.linkedin.com/in/agam-jain-5a8b95151/","✅ DIRECT FOUNDER EMAIL — YC W24; serverless GPU infra"),
    ("Tensorfuse","Samagra Sharma","Co-Founder & CEO","samagra@tensorfuse.io","https://in.linkedin.com/in/samagra-sharma-4476bb135","✅ DIRECT CEO EMAIL — YC W24; also founders@tensorfuse.io"),
    ("GoSats","Mohammed Roshan","Co-Founder & CEO","roshan@gosats.io","https://in.linkedin.com/in/roshanaslam","✅ DIRECT CEO EMAIL — YC W22; Bitcoin rewards; Bengaluru"),
    ("Emergent","Mukund Jha","Co-Founder & CEO","team@emergent.sh","https://www.linkedin.com/in/mukundjha/","YC S24; $100M raised; ex-Dunzo CTO (Google-backed); twin brother Madhav Jha is CTO"),
    ("Flagright","Madhu G Nadig","Co-Founder & CTO","gdpr@flagright.com","https://www.linkedin.com/in/madhugnadig/","YC W22; $4.3M seed; AML compliance; Bangalore office; CEO Baran Ozkan based in SF"),
    # ── Round 18: More Bangalore founder contacts ───────────────────────────
    ("Rivia.AI","Samay Jain","Co-Founder & CEO","samay@rivia.ai","https://www.linkedin.com/in/samayjain/","✅ DIRECT CEO EMAIL — YC S21; 5 employees; interactive product demos; HIRING"),
    ("Infinity","Sourav Choraria","Co-Founder & CEO","sourav@infinityapp.in","https://www.linkedin.com/in/souravchoraria/","YC W24; $1.9M pre-seed; cross-border fintech; brother Sidharth is co-founder"),
    ("Vahan.ai","Madhav Krishna","Founder & CEO","madhav@vahan.ai","https://in.linkedin.com/in/madhavkrishna","✅ DIRECT CEO EMAIL — $23.7M funded; Columbia CS; WEF Tech Pioneer; Khosla + Founders Fund backed"),
    ("smallest.ai","Sudarshan Kamath","Co-Founder & CEO","info@smallest.ai","https://www.linkedin.com/in/sudarshankamath/","Indiranagar Bangalore; voice AI; ex-Bosch AI; viral hiring post; Co-founder: Akshat Mandloi"),
    ("Strac","Aatish Mandelecha","Founder & CEO","aatish@strac.io","https://www.linkedin.com/in/aatishmandelecha/","YC W22; ex-Amazon 11 yrs payments; DLP for SaaS/Cloud/GenAI; Bengaluru office"),
    ("Kula","Achuthanand Ravi","Co-Founder & CEO","hello@kula.ai","https://www.linkedin.com/in/achuthanand-ravi/","$15M seed; ex-founding recruiter Freshworks/Uber/Stripe; AI-native ATS"),
    # ── Round 19: More Bangalore founder contacts ───────────────────────────
    ("Sarvam AI","Pratyush Kumar","Co-Founder & CEO","careers@sarvam.ai","https://www.linkedin.com/in/pratyush-kumar-8844a8a/","India's sovereign AI; govt IndiaAI Mission contract; IIT Bombay PhD; ex-Google Brain; 29 open roles"),
    ("Locale.ai","Aditi Sinha","Co-Founder & CEO","aditi@locale.ai","https://www.linkedin.com/in/aditisinha1002/","✅ DIRECT CEO EMAIL — Forbes 30U30; $5.5M funded; geospatial analytics"),
    ("FamApp","Sambhav Jain","Co-Founder & CEO","sambhav@fampay.in","https://in.linkedin.com/in/sambhavanandjain","YC S19; Forbes 30U30; IIT Roorkee; fintech for teens; also Kush Taneja co-founder"),
    ("Rocketium","Satej Sirur","Co-Founder & CEO","satej@rocketium.com","https://www.linkedin.com/in/satejsirur/","✅ DIRECT CEO EMAIL — AI creative automation; Blume Ventures; RETHINK Retail Top AI Leader"),
    ("Rigi","Swapnil Saurav","Co-Founder & CEO","swapnil@rigi.club","https://www.linkedin.com/in/linkswapnil/","$25M funded; Elevation Capital; HSR Layout; prev founded HalaPlay (acquired by Nazara)"),
    # ── Round 20: DevTools / SaaS / Open Source founder contacts ────────────
    ("SigNoz","Pranay Prateek","Co-Founder & CEO","pranay@signoz.io","https://www.linkedin.com/in/pranay01/","✅ DIRECT CEO EMAIL — YC W21; open source observability; also hiring@signoz.io"),
    ("Clarisights","Arun Srinivasan","Co-Founder & CEO","arun@clarisights.com","https://www.linkedin.com/in/arun-srinivasan-clarisights/","✅ DIRECT CEO EMAIL — Indiranagar Bangalore; marketing analytics for enterprises"),
    ("Dukaan","Suumit Shah","Co-Founder & CEO","suumit@mydukaan.io","https://in.linkedin.com/in/suumitshah","Bengaluru; DIY e-commerce; CTO Subhash Choudhary; also support@mydukaan.io"),
    ("Scribble Data","Dr. Venkata Pingali","Co-Founder & CEO","venkata@scribbledata.io","https://in.linkedin.com/in/pingali","✅ DIRECT CEO EMAIL — IIT Bombay; MLOps; Blume Ventures; Bangalore+Toronto; HIRING"),
    ("Sprinto","Girish Redekar","Co-Founder & CEO","girish@sprinto.com","https://www.linkedin.com/in/girishredekar/","$32.2M Series B; autonomous compliance; prev RecruiterBox (acquired by Turn/River Capital)"),
    # ── Round 21: IT services + AI dev founder contacts ─────────────────────
    ("Krazimo","Akhil Verghese","CEO","akhil@krazimo.ai","https://www.linkedin.com/in/akhilverghese/","✅ DIRECT CEO EMAIL — ex-Google; AI engineering; Bellandur/HSR Bangalore"),
    ("Reckonsys","Sathish Visanagiri","Founder & CEO","sathish@reckonsys.com","https://www.linkedin.com/in/sathish-visanagiri/","Sarjapur Road Bangalore; AI chatbots; custom software; also info@reckonsys.com"),
    ("Pace Wisdom","Bharath Jatangi","Co-Founder","contact@pacewisdom.com","","Rajajinagar Bangalore; AI product dev; Co-founder: Mohan Thimmadasaiah"),
    ("Evnek Technologies","Ashis Kumar Sahoo","Director","info@evnek.com","","Whitefield Bangalore; GenAI/LLM/Cloud; also directors: Debasish Panda, Smita Sahoo"),
    # ── Round 22: Cybersecurity + InsurTech founder contacts ────────────────
    ("CloudSEK","Rahul Sasi","Co-Founder & CEO","rahul.sasi@cloudsek.com","https://www.linkedin.com/in/fb1h2s/","✅ DIRECT CEO EMAIL — $19M Series B1; ethical hacker; cybersecurity; 32 open positions; also careers@cloudsek.com"),
    ("Plum","Abhishek Poddar","Co-Founder & CEO","abhishek@plumhq.com","https://in.linkedin.com/in/abhishek24","✅ DIRECT CEO EMAIL — $36M Series B; InsurTech; employee health; Co-founder: Saurabh Arora"),
    ("Nova Benefits","Saransh Garg","Co-Founder & CEO","admin@getnovaapp.com","https://www.linkedin.com/in/saransh-garg/","Bangalore; employee wellness SaaS; Co-founder: Yash Gupta; Ph: +91 91673 39156"),
    # ── Round 23: AI Voice / LegalTech / EdTech / HealthTech / Deep Tech founders ──
    ("Myelin Foundry","Gopichand Katragadda","Founder & CEO","social@myelinfoundry.com","https://in.linkedin.com/in/gkatragadda","Whitefield Bangalore; ex-Group CTO Tata Sons; AI edge devices; Ph: +91 80 6190 4242"),
    ("SpotDraft","Shashank Bijapur","Co-Founder & CEO","shashank@spotdraft.com","https://in.linkedin.com/in/shashankbijapur","✅ DIRECT CEO EMAIL — HSR Layout Bangalore; $113M+ funded; AI contract management; 234 employees"),
    ("Murf AI","Ankur Edkie","Co-Founder & CEO","ankur@murf.ai","https://www.linkedin.com/in/ankuredkie/","✅ DIRECT CEO EMAIL — HSR Layout Bangalore; $11.5M Series A; AI voice/TTS; 6M+ users; IIT-KGP; ex-Goldman"),
    ("Presentations.AI","Sumanth Raghavendra","Co-Founder & CEO","sumanth@presentations.ai","https://in.linkedin.com/in/raghavendrasumanth","✅ DIRECT CEO EMAIL — Jayanagar Bangalore; Accel-backed; 10M+ users; AI presentations"),
    ("SuperKalam","Vimal Singh Rathore","CEO","vimal@kalam.in","https://in.linkedin.com/in/vimal-rathore","✅ DIRECT CEO EMAIL — YC W23; $2M seed; AI edtech; 11 open roles; hiring: join@superkalam.com"),
    ("Richpanel","Amit RG","CEO & Founder","amit@richpanel.com","https://www.linkedin.com/in/amit-rg","YC + Sequoia; 2000+ brands; AI customer support; HIRING ML Engineer in Bangalore"),
    ("Even Healthcare","Mayank Banerjee","Co-Founder & CEO","mayank@even.in","https://in.linkedin.com/in/mayank-banerjee-b081507b","✅ DIRECT CEO EMAIL — Indiranagar Bangalore; $30M Series A Khosla; healthtech; also careers@even.in"),
    # ── Round 24: B2B Marketing AI / Voice AI / AgriTech founder contacts ─────
    ("Factors.ai","Srikrishna Swaminathan","Co-Founder & CEO","srikrishna@factors.ai","https://www.linkedin.com/in/srifactorsai/","✅ DIRECT CEO EMAIL — Elevation Capital; ex-InMobi VP; also CPO Praveen Das praveen@factors.ai"),
    ("Bolna AI","Maitreya Wagh","Co-Founder & CEO","maitreya@bolna.ai","https://www.linkedin.com/in/maitreya-wagh/","YC F25; $6.3M seed General Catalyst; IIT Delhi + ex-Bain; voice AI for Indian languages"),
    ("Fasal","Shailendra Tiwari","Founder & CEO","connect@wolkus.com","https://in.linkedin.com/in/shailendra-tiwari-fasal","$19.4M funded; agritech AI/IoT; 12K+ farmers; general contact email"),
    ("Skit.ai","Sourabh Gupta","Co-Founder & CEO","scale@skit.ai","https://www.linkedin.com/in/sourabhsg/","Forbes 30U30 Asia; voice AI; hiring email scale@skit.ai for Bangalore roles; IIT Roorkee"),
    # ── Round 25: AI CX / Inference Infra / Marketing Cloud founder contacts ──
    ("Hiver","Niraj Ranjan Rout","Founder & CEO","niraj@hiverhq.com","https://www.linkedin.com/in/nirajranjan/","HSR Layout Bangalore; $46.2M funded; AI helpdesk for Google Workspace; IIT KGP; 10K+ teams"),
    ("Simplismart","Amritanshu Jain","Co-Founder & CEO","amritanshu@simplismart.ai","https://www.linkedin.com/in/jainamritanshu/","Richmond Town Bangalore; $14M funded Accel; AI inference engine; BITS Pilani; ex-Oracle"),
    ("Pixis","Shubham A Mishra","Co-Founder & Global CEO","shubham@pixis.ai","https://www.linkedin.com/in/shubhammishra01/","Bengaluru; codeless AI marketing infra; BITS Pilani; prev Absentia VR"),
    # ── Round 26: Digital Adoption / Open Source DevTools founder contacts ────
    ("Whatfix","Khadim Batti","Co-Founder & CEO","khadim@whatfix.com","https://in.linkedin.com/in/khadim","Bengaluru HQ; 800+ employees; digital adoption platform; ZoomInfo k***@whatfix.com confirmed"),
    ("ToolJet","Navaneeth Padanna Kalathil","Founder & CEO","navaneeth@tooljet.com","https://www.linkedin.com/in/navaneeth-pk/","✅ DIRECT CEO EMAIL — open source low-code; HIRING engineers/PMs/designers; Ph: +91 9400812423"),
    ("Hasura (PromptQL)","Tanmai Gopal","Co-Founder & CEO","tanmai@hasura.io","https://www.linkedin.com/in/tanmaig/","Bangalore + SF; GraphQL + PromptQL AI; IIT Madras; unicorn; hiring via hasura.io/careers"),
    # ── Round 27: SpaceTech + Gaming founder contacts ─────────────────────────
    ("Pixxel","Awais Ahmed","Founder & CEO","awais@pixxel.space","https://in.linkedin.com/in/awaisahmedna","Bengaluru; $96M; hyperspectral satellites; BITS Pilani; 177 employees"),
    ("GalaxEye","Suyash Singh","Founder & CEO","suyash@galaxeye.space","https://in.linkedin.com/in/suyashaoc","🎯 SEEKING INTERNS — Bengaluru; IIT Madras spin-off; multi-sensor imaging; careers.galaxeye.space"),
    ("Digantara","Anirudh Sharma","Co-Founder & CEO","anirudh@digantara.co.in","https://in.linkedin.com/in/anirudh-sharma-digantara","✅ DIRECT CEO EMAIL — Hebbal Bangalore; $50M Dec 2025; space surveillance; HIRING"),
    ("Bombay Play","Sruthi (Hiring)","Recruiting","sruthi@bombayplay.com","https://www.linkedin.com/company/bombay-play","✅ DIRECT HIRING EMAIL — Indiranagar Bangalore; casual games; CEO Oliver Jones"),
    ("Mayhem Studios","Ojas Vipat","Founder & CEO","careers@mayhem-studios.com","https://www.linkedin.com/in/ojasvipat/","Bangalore; AAA mobile gaming; MPL-backed; Underworld Gang Wars"),
    # ── Round 28: EV / Mobility founder contacts ──────────────────────────────
    ("Ultraviolette","Narayan Subramaniam","Co-Founder & CEO","narayan@ultraviolette.com","https://in.linkedin.com/in/narayan-s","Domlur Bangalore; F77 e-motorcycle; TVS-backed; estimated email pattern"),
    ("River Mobility","Aravind Mani","Co-Founder & CEO","aravind@rideriver.com","https://in.linkedin.com/in/aravindmani","Bengaluru; Yamaha-backed e-scooters; ZoomInfo confirmed a***@rideriver.com"),
    ("Simple Energy","Suhas Rajkumar","Founder & CEO","suhas@simpleenergy.in","https://in.linkedin.com/in/suhas-rajkumar-277824150","Yelahanka Bangalore; ₹250Cr raised; Simple One scooter; estimated email pattern"),
    # ── Round 29: Emotion AI / Conversational AI / Sovereign LLM founders ─────
    ("Entropik","Ranjan Kumar","Co-Founder & CEO","ranjan@entropiktech.com","https://www.linkedin.com/in/ranjan-kr/","Bengaluru; Series B Emotion AI; ZoomInfo confirmed R***@entropiktech.com; Ph: 080-4375-9863"),
    ("Senseforth.ai","Shridhar Marri","Co-Founder & CEO","shridhar@senseforth.ai","https://in.linkedin.com/in/shridharmarri","Girinagar Bangalore; enterprise conversational AI; email pattern estimated"),
    ("CoRover.ai","Ankush Sabharwal","Founder & CEO","ankush@corover.ai","https://in.linkedin.com/in/ankushsabharwal","Bangalore; BharatGPT creator; 1B+ users via IRCTC; email pattern estimated"),
    # ── Round 30: HR Tech / Developer Hiring / CX Analytics founders ──────────
    ("Springworks","Kartik Mandaville","Founder & CEO","kartik@springworks.in","https://www.linkedin.com/in/kartik-mandaville-springworks/","Bengaluru; remote-first ALWAYS HIRING; SpringVerify/SpringRecruit; CMU grad"),
    ("HackerEarth","Vikas Aditya","CEO","support@hackerearth.com","https://www.linkedin.com/in/vikasaditya/","Bengaluru; dev assessment platform; IIT Roorkee founders; 7M+ dev community"),
    ("Clootrack","Shameel Abdulla","Co-Founder & CEO","shameel@clootrack.com","https://www.linkedin.com/in/shameelabdulla/","Bangalore; $4M AI CX analytics; Salesforce Ventures; email pattern estimated"),
    # ── Round 31: Sales AI / WealthTech founders ──────────────────────────────
    ("Salesken","Surga Thilakan","Co-Founder & CEO","surga@salesken.ai","https://in.linkedin.com/in/surga-thilakan-0196994","✅ DIRECT CEO EMAIL — Bengaluru; $41M Microsoft+Sequoia; sales AI; ex-Goldman; Ph: +91 80416 49503"),
    ("Smallcase","Vasanth Kamath","Founder & CEO","vasanth@smallcase.com","https://in.linkedin.com/in/vasanthskamath","Richmond Road Bangalore; wealthtech; IIT-KGP; 367 employees; email pattern estimated"),
    # ── Round 32: Gandhinagar second pass — web tech HR contacts ──────────────
    ("BOSC Tech Labs","HR Team","HR","hr@bosctechlabs.com","https://in.linkedin.com/company/bosc-tech-labs","✅ DIRECT HR EMAIL — Sargasan Gandhinagar; 5 open jobs; Flutter/React"),
    ("Samcom Technologies","Ketan Jadhav","Founder & CEO","ketan@samcomtechnologies.com","https://theorg.com/org/samcom-technologies/org-chart/ketan-jadhav","Motera/Gandhinagar border; AI automation; ZoomInfo k***@samcomtechnologies.com"),
    ("Evolvision Technologies","Jimit Joshi","Founder & CEO","business@evolvision.com","https://in.linkedin.com/in/jimitjoshi","✅ PHONE: +91-9510645454 — Kudasan Gandhinagar"),
    ("Unity Infoway","Rahul Gondaliya","Founder","info@unityinfoway.com","https://in.linkedin.com/company/unity-infoway","Kudasan Gandhinagar; web/mobile dev since 2011"),
]

headers = ["#","Company","Website","Category","General Contact Email","Email Type","HR Person Name","HR Person Title","HR Person Direct Email","LinkedIn Profile","Priority","Notes"]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", start_color="1F3864")
DATA_FONT   = Font(name="Arial", size=9)
ALT_FILL    = PatternFill("solid", start_color="DCE6F1")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")
GREEN_FILL  = PatternFill("solid", start_color="C6EFCE")
YELLOW_FILL = PatternFill("solid", start_color="FFEB9C")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP   = Alignment(wrap_text=True, vertical="top")
thin   = Side(style="thin", color="B8CCE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws):
    for cell in ws[1]:
        cell.font=HEADER_FONT; cell.fill=HEADER_FILL
        cell.alignment=CENTER; cell.border=BORDER

def style_row(ws, r, alt):
    bg = ALT_FILL if alt else WHITE_FILL
    for cell in ws[r]:
        cell.font=DATA_FONT; cell.fill=bg
        cell.border=BORDER; cell.alignment=WRAP

def colour_priority(cell, val):
    if val=="High":
        cell.fill=GREEN_FILL
        cell.font=Font(name="Arial",size=9,bold=True,color="276221")
    elif val=="Medium":
        cell.fill=YELLOW_FILL
        cell.font=Font(name="Arial",size=9,bold=True,color="9C5700")

col_widths=[5,22,22,14,28,10,22,22,30,42,10,40]

def setup_sheet(ws):
    ws.append(headers); style_header(ws); ws.freeze_panes="A2"
    for idx,w in enumerate(col_widths,1):
        ws.column_dimensions[get_column_letter(idx)].width=w

# Sheet 1
ws1=wb.active; ws1.title="All Companies"; setup_sheet(ws1)
for i,row in enumerate(companies,2):
    ws1.append(list(row)); style_row(ws1,i,i%2==0)
    colour_priority(ws1.cell(row=i,column=11),row[10])

# Sheet 2
ws2=wb.create_sheet("Priority Dubai Companies"); setup_sheet(ws2)
r=2
for row in companies:
    if row[10]=="High":
        ws2.append(list(row)); style_row(ws2,r,r%2==0)
        colour_priority(ws2.cell(row=r,column=11),row[10]); r+=1

# Sheet 3
ws3=wb.create_sheet("Named HR Contacts")
hr_h=["Company","HR Person Name","HR Person Title","HR Person Direct Email","LinkedIn Profile","Action"]
ws3.append(hr_h); style_header(ws3); ws3.freeze_panes="A2"
for idx,w in enumerate([18,22,28,35,55,55],1):
    ws3.column_dimensions[get_column_letter(idx)].width=w
for i,row in enumerate(hr_contacts,2):
    ws3.append(list(row)); style_row(ws3,i,i%2==0)
    if "VERIFIED" in row[5]:
        ws3.cell(row=i,column=4).fill=GREEN_FILL
        ws3.cell(row=i,column=4).font=Font(name="Arial",size=9,bold=True,color="276221")

# Sheet 4
ws4=wb.create_sheet("Email Template")
ws4.column_dimensions["A"].width=18; ws4.column_dimensions["B"].width=85
ws4.append(["DUBAI INTERNSHIP OUTREACH - EMAIL TEMPLATE",""])
ws4.merge_cells("A1:B1")
ws4["A1"].font=Font(name="Arial",bold=True,size=13,color="FFFFFF")
ws4["A1"].fill=PatternFill("solid",start_color="1F3864")
ws4["A1"].alignment=Alignment(horizontal="center",vertical="center")
ws4.row_dimensions[1].height=24
template=[
    ("SUBJECT","Paid Internship Enquiry - Eesh Saxena (CS Undergrad)"),
    ("",""),
    ("BODY","Good morning,"),
    ("",""),
    ("","My name is Eesh Saxena. I wanted to reach out and enquire if you have any paid internship opportunities available."),
    ("",""),
    ("","I've completed research internships at prestigious institutions, built production web apps with the MERN stack, and worked on ML projects ranging from NLP and computer vision to edge AI deployment. I'm comfortable moving between frontend, backend, and model training depending on what's needed."),
    ("",""),
    ("","I've attached my resume for your reference. Would love to hear from you if there's a fit."),
    ("",""),
    ("","Best,"),
    ("","Eesh Saxena"),
    ("","7976212108 | eeshsaxena@gmail.com"),
]
for i,(lbl,val) in enumerate(template,2):
    ws4.cell(row=i,column=1).value=lbl
    ws4.cell(row=i,column=2).value=val
    ws4.cell(row=i,column=1).font=Font(name="Arial",bold=True,size=9,color="1F3864")
    ws4.cell(row=i,column=2).font=Font(name="Arial",size=10)
    ws4.cell(row=i,column=2).alignment=Alignment(wrap_text=True,vertical="top")
    ws4.row_dimensions[i].height=35 if val and len(val)>80 else 15

try:
    wb.save(r"C:\Users\eeshs\Downloads\Dubai_Internship_Outreach.xlsx")
except PermissionError:
    wb.save(r"C:\Users\eeshs\Downloads\Dubai_Internship_Outreach_v2.xlsx")
    print("Saved as v2 (original was open)")
print("DONE")
