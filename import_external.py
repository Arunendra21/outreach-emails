import csv
import os

# Imports external contact files into extra_contacts.csv (same schema as hr_contacts.csv).
# gen_batches.py merges extra_contacts.csv into the sendable batch files.

SRC = r"D:\New folder\python-3.12.3-amd64"

LINKEDIN_CSVS = [
    "jaipur.csv",
    "Multiple_Folders_2026-04-06 (1).csv",
    "Multiple_Folders_2026-04-06.csv",
]
GROWTH_XLSX = "growth-list-free-forever-march-2026.xlsx"  # (1) copy is identical

out = []
seen = set()

def add(company, name, title, email, linkedin, notes):
    email = (email or "").strip()
    if "@" not in email or " " in email:
        return
    key = email.lower()
    if key in seen:
        return
    seen.add(key)
    out.append([company or "—", name or "—", title or "—", email, linkedin or "", notes])

# LinkedIn people exports
for fname in LINKEDIN_CSVS:
    path = os.path.join(SRC, fname)
    if not os.path.exists(path):
        print("MISSING:", path)
        continue
    with open(path, encoding="utf-8", errors="replace") as f:
        for r in csv.DictReader(f):
            name = " ".join(x for x in [r.get("First Name", ""), r.get("Last Name", "")] if x).strip()
            work = r.get("Work Email", "").strip()
            personal = r.get("Personal Email", "").strip()
            extras = [r.get(k, "").strip() for k in ("Additional Email 1", "Additional Email 2", "Additional Email 3")]
            emails = [e for e in [work, personal] + extras if "@" in e]
            if not emails:
                continue
            primary, alts = emails[0], emails[1:]
            notes_bits = []
            if alts:
                notes_bits.append("alt: " + ", ".join(alts))
            for pk in ("Phone", "Phone 2"):
                if r.get(pk, "").strip():
                    notes_bits.append("Ph: " + r[pk].strip())
            loc = r.get("Location", "").split(",")[0].strip()
            if loc:
                notes_bits.append(loc)
            add(r.get("Company", ""), name, r.get("Title", ""), primary,
                r.get("Linkedin URL", ""), "; ".join(notes_bits))

# Growth List funded companies (verified contact emails)
from openpyxl import load_workbook
gpath = os.path.join(SRC, GROWTH_XLSX)
if os.path.exists(gpath):
    ws = load_workbook(gpath, read_only=True).worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr = {str(v): i for i, v in enumerate(rows[0])}
    for r in rows[1:]:
        if not r or not r[hdr["Contact Email"]]:
            continue
        status = str(r[hdr["Email Status"]] or "")
        funding = str(r[hdr["Funding Amount (in USD)"]] or "")
        notes = "; ".join(x for x in [
            "Growth List Mar-2026 funded co",
            f"{r[hdr['City']]}, {r[hdr['Country']]}",
            f"raised ${funding}" if funding not in ("", "unknown", "None") else "",
            str(r[hdr["Funding Type"]] or ""),
            f"email {status}",
        ] if x)
        add(str(r[hdr["Name"]] or ""), "Team", "Contact", str(r[hdr["Contact Email"]]),
            str(r[hdr["LinkedIn"]] or ""), notes)

with open("extra_contacts.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["Company", "Name", "Title", "Email", "LinkedIn", "Action"])
    w.writerows(out)

print(f"extra_contacts.csv: {len(out)} contacts")
